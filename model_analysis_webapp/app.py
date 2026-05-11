from flask import Flask, render_template, request, send_from_directory
import os
import pandas as pd
import openpyxl
from werkzeug.utils import secure_filename
import uuid
import re
import datetime
from openpyxl.utils import get_column_letter

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

# Hardcoded variables
def get_hardcoded_vars(sheet_name):
    model_vars = ["Net Sales", "Depreciation Expense", "Amortization", "Cost of Goods Sold", "COGS(excl Dep&Amort) % of Sales",
                  "Gross Profit", "Gross Profit Margin", "Selling, Gen & Admin Expenses", "Research & Development Expenses",
                  "% of Net Sales", "Operating Income", "Operating EBITDA", "Interest Income", "Interest Expense",
                  "Comprehensive Financing (Net)", "Other Income/(expenses)", "Earnings Before Taxes", "Income Taxes",
                  "Minority Interest", "Income Before Extra Items", "Net Income", "Proforma Adjustments", "Proforma Net Income",
                  "Shares Outstanding (mm) (Fully Diluted)", "EPS (Fully diluted)", "DPS", "Total Current Assets",
                  "Net Property Plant and Equip", "Total Other Long-Term Assets", "Total Assets", "Total Current Liabilities",
                  "Total Long-Term Liabilities", "Total Common Equity", "Total Liabs and Equity", "Cash Flows from Operations",
                  "Cash Flows from Investing", "Cash Flow from Financing", "Net Working Capital (NWC)", "Unlevered FCFF"]

    ratio_vars = ["ROE", "ROA", "ROCE", "ROIC", "Gross Profit Margin", "Operating Profit Margin", "EBITDA Margin",
                  "Pre-Tax Margin", "Net Margin", "CF", "FCF", "Days' Receivables", "Days' Inventory", "Days' Payable",
                  "Cash Conversion Cycle", "Receivables Turnover", "Inventory Turnover", "Current Ratio", "Quick Ratio",
                  "R&D / Net Sales", "Working Capital / Net Sales", "Capex / Net Sales", "Capex / Depreciation",
                  "Depreciation / Gross P, P& E", "Depreciation / Net Sales", "Operating Leverage"]

    dcf_vars = ["Cost of Equity", "After-Tax Cost of Debt", "WACC (Discount Rate)", "Unlevered FCFF"]

    if sheet_name == "Model":
        return model_vars
    elif sheet_name == "Ratio Analysis":
        return ratio_vars
    elif sheet_name == "DCF":
        return dcf_vars
    return []


def get_row_description(ws, row):
    """
    Try to get a non-empty description from columns A, B, or C for a given row
    """
    for col_letter in ['A', 'B', 'C', 'D']:
        cell = ws[f"{col_letter}{row}"]
        val = cell.value
        if val is not None and str(val).strip() != "":
            return str(val).strip()
    return ""



# Period extraction

def extract_periods(ws, max_row=10):
    pattern = re.compile(r'\b(\d{1,2}Q\d{2,4}E|FY\d{2,4}E|\d{4}E)\b', re.I)
    seen = set()
    ordered_periods = []

    for row in range(1, min(ws.max_row, max_row) + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            val = cell.value

            if val is None:
                continue

            # Only match explicitly formatted strings ending with "E"
            if isinstance(val, str):
                val = val.strip()
                match = pattern.fullmatch(val)
                if match:
                    period = match.group(1)
                    if period not in seen:
                        seen.add(period)
                        ordered_periods.append(period)

    return ordered_periods



# Extract dropdown options from Excel

def extract_dropdown_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_sheets = wb.sheetnames

    exclude = {"Desc", "Fundamental", "ErrorChk", "LinkC", "Street Est", "Price Data"}
    custom_order = ["Model", "Ratio Analysis", "DCF", "SOP"]

    ordered_sheets = [s for s in custom_order if s in all_sheets and s not in exclude]
    remaining_sheets = sorted([s for s in all_sheets if s not in ordered_sheets and s not in exclude])
    sheet_options = ordered_sheets + remaining_sheets

    sheet_data = {}

    for sheet in sheet_options:
        ws = wb[sheet]
        hardcoded = get_hardcoded_vars(sheet)
        used_vars = set(hardcoded)
        var_set = set()

        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = str(row[0]).strip() if row[0] is not None else ""
            if val and val not in used_vars:
                var_set.add(val)

        variables = hardcoded + sorted(var_set)
        periods = extract_periods(ws)
        sheet_data[sheet] = {"variables": variables, "periods": periods}

    return sheet_options, sheet_data




@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'workbook' in request.files:
            file = request.files['workbook']
            filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(file_path)

            sheet_options, sheet_data = extract_dropdown_data(file_path)

            return render_template('index.html',
                                   file_uploaded=True,
                                   file_path=filename,
                                   uploaded_filename=file.filename,
                                   sheet_options=sheet_options,
                                   sheet_data=sheet_data)

        else:
            form = request.form
            file_path = os.path.join(UPLOAD_FOLDER, form['file_path'])
            sheet_options, sheet_data = extract_dropdown_data(file_path)

            if 'sheet' in form and ('variable' not in form or 'period' not in form):
                return render_template('index.html',
                                       file_uploaded=True,
                                       file_path=form['file_path'],
                                       uploaded_filename=form.get('uploaded_filename', ''),
                                       sheet_options=sheet_options,
                                       sheet_data=sheet_data,
                                       selected_sheet=form['sheet'])

            sheet = form['sheet']
            variable = form['variable']
            period = form['period']

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet]

            target_row = None
            for row in ws.iter_rows(min_col=1, max_col=1):
                raw = row[0].value
                val = str(raw).strip() if raw is not None else ""
                if val == variable:
                    target_row = row[0].row
                    break

            if not target_row:
                return "Variable not found", 400

            target_col = None
            for r in range(1, 11):
                for c in range(1, ws.max_column + 1):
                    raw = ws.cell(r, c).value
                    val = str(raw).strip() if raw is not None else ""
                    if val == period:
                        target_col = c
                        break
                if target_col:
                    break

            if not target_col:
                return "Period not found", 400
            



            forecasted_value = ws.cell(target_row, target_col).value
            description = get_row_description(ws, target_row)
            cell_ref = f"{get_column_letter(target_col)}{target_row}"

            actual_historical_value = "N/A"
            if period.endswith("E"):
                is_q = 'Q' in period
                is_h = 'H' in period

                for c in range(target_col - 1, 0, -1):
                    for r in range(1, 11):
                        header = ws.cell(r, c).value
                        if not header:
                            continue
                        header_str = str(header).strip()
                        if header_str.endswith("E"):
                            continue

                        if (not is_q and not is_h and 'Q' not in header_str and 'H' not in header_str and
                            '-' not in header_str and ':' not in header_str and
                            (header_str.isdigit() or header_str.startswith('FY'))):
                            actual_historical_value = ws.cell(target_row, c).value
                            break
                        elif is_q and 'Q' in header_str:
                            actual_historical_value = ws.cell(target_row, c).value
                            break
                        elif is_h and 'H' in header_str:
                            actual_historical_value = ws.cell(target_row, c).value
                            break
                    if actual_historical_value != "N/A":
                        break

            # Validate values
            forecasted_valid = forecasted_value is not None and isinstance(forecasted_value, (int, float))
            actual_valid = actual_historical_value not in ["", "N/A", None]

            # Warning block moved BEFORE file return
            if not forecasted_valid or not actual_valid:
                warnings = []
                if not forecasted_valid:
                    warnings.append("⚠ Forecasted value is missing or not a number.")
                if not actual_valid:
                    warnings.append("⚠ Actual historical value is missing or invalid.")
                
                return render_template('index.html',
                                    file_uploaded=True,
                                    file_path=form['file_path'],
                                    uploaded_filename=form.get('uploaded_filename', ''),
                                    sheet_options=sheet_options,
                                    sheet_data=sheet_data,
                                    selected_sheet=sheet,
                                    warning=" ".join(warnings))

            # If all good, generate CSV and return
            df = pd.DataFrame({
                "Sheet": [sheet],
                "Variable": [variable],
                "Period": [period],
                "Forecasted Value": [forecasted_value if forecasted_valid else ""],
                "Forecasted Cell Reference": [cell_ref],
                "Actual Historical Value": [actual_historical_value if actual_valid else ""],
                "Row Description": [description]
            })

            result_filename = f"result_{uuid.uuid4().hex}.csv"
            result_file = os.path.join(RESULT_FOLDER, result_filename)
            df.to_csv(result_file, index=False)

            return send_from_directory(RESULT_FOLDER, result_filename, as_attachment=True)


    return render_template('index.html')




if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

