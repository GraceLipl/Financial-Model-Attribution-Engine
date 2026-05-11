
import datetime
def build_column_period_map(wb_values, sheet_name, start_row=1, today_date=None, max_row=150):
    """
    Build column to period map (automated) - EXACTLY as in your Jupyter notebook
    """
    if today_date is None:
        today_date = datetime.date.today()

    sheet = wb_values[sheet_name]  # Access worksheet by name from Workbook
    col_to_period = {}

    # Enhanced pattern to match more period formats - same as lookup_column_period
    pattern = re.compile(r'\b(\d{1,2}Q\d{2,4}E?|\d{4}E?|FY\d{2,4}E?)\b', re.I)

    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        period = ""

        for row in range(start_row, min(max_row + 1, sheet.max_row + 1)):
            cell_obj = sheet[f"{col_letter}{row}"]
            cell_value = cell_obj.value

            if cell_value is None:
                continue

            # Process string values - look for period patterns
            if isinstance(cell_value, str):
                text = cell_value.strip()
                
                # Search for period pattern in the text
                match = pattern.search(text)
                if match:
                    period = match.group(1)  # Return the matched period
                    break

            # Process numeric values - handle years stored as numbers
            elif isinstance(cell_value, (int, float)):
                # Check if it's a 4-digit year
                if 2000 <= cell_value <= 2050:
                    year = int(cell_value)
                    # Determine if it's forecast based on year threshold
                    forecast_threshold = 2025  # Years >= 2025 are considered forecast
                    if year >= forecast_threshold:
                        period = f"{year}E"
                    else:
                        period = str(year)
                    break

            # Skip other types (dates, etc.)

        if period:
            col_to_period[col_letter] = period

    return col_to_period

def build_all_sheets_col_to_period(wb_values, today_date=None, max_row=150):
    """
    Build col_to_period maps for all sheets in the workbook - EXACTLY as in your Jupyter notebook
    """
    all_col_to_period = {}
    
    for sheet_name in wb_values.sheetnames:  # Use .sheetnames for openpyxl Workbook
        try:
            col_to_period = build_column_period_map(
                wb_values, sheet_name, start_row=1, today_date=today_date, max_row=max_row
            )
            if col_to_period:  # Only store if we found periods
                all_col_to_period[sheet_name] = col_to_period
                print(f"✓ {sheet_name}: Found {len(col_to_period)} period columns")
            else:
                print(f"⚠ {sheet_name}: No periods found")
        except Exception as e:
            print(f"✗ {sheet_name}: Error - {e}")
    
    return all_col_to_period

def lookup_column_period_from_map(all_col_to_period, sheet_name, cell_ref):
    """
    Helper function to get period for a specific cell from the all_col_to_period map
    """
    if sheet_name not in all_col_to_period:
        return ""
    
    col_letter = re.match(r"[A-Z]+", cell_ref).group(0)
    return all_col_to_period[sheet_name].get(col_letter, "")


import re
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils import range_boundaries

def parse_sheet_reference(ref):
    """
    Parse a sheet reference like 'Sheet Name'!A1 or Sheet!A1
    Returns (sheet_name, cell_ref)
    """
    if "!" not in ref:
        return None, ref
    
    # Split by the last occurrence of ! to handle cases like 'Sheet!Name'!A1
    parts = ref.rsplit("!", 1)
    sheet_part = parts[0]
    cell_part = parts[1]
    
    # Remove quotes if present
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1]
    
    return sheet_part, cell_part

def resolve_label_with_hierarchy(wb_values, sheet_name, row):
    """
    Get cell description from columns A, B, C without bold header enrichment
    Modified to work with wb_values parameter and sheet_name
    """
    #print(f"DEBUG: Getting description for row {row}")
    
    # Get the sheet object from wb_values
    sheet = wb_values[sheet_name]
    
    # Check columns A, B, C in order to find the first non-empty cell
    for col_letter in ['A', 'B', 'C']:
        cell = sheet[f"{col_letter}{row}"]
        #print(f"DEBUG: Checking {col_letter}{row} = {cell.value}")
        
        if cell.value is not None:
            cell_value = cell.value
            
            # If cell has a formula, follow it
            if hasattr(cell, 'data_type') and (cell.data_type == "f" or (isinstance(cell_value, str) and cell_value.startswith("="))):
                #print(f"DEBUG: Found formula in {col_letter}{row}: {cell_value}")
                m = re.match(r"=([^!]+)!([A-Z]+[0-9]+)", str(cell_value))
                if m:
                    target_sheet, target_cell = m.groups()
                    try:
                        # Use the wb_values parameter instead of global wb_values
                        cell_value = wb_values[target_sheet][target_cell].value
                        #print(f"DEBUG: Formula resolved to: {cell_value}")
                        if cell_value is None:
                            cell_value = ""
                        else:
                            cell_value = str(cell_value).strip()
                    except:
                        cell_value = ""
            
            result = str(cell_value).strip() if cell_value else ""
            #print(f"DEBUG: Final result for row {row}: '{result}'")
            return result
    
    # If no cell found, return row number
    #print(f"DEBUG: No cell found for row {row}, returning Row {row}")
    return f"Row {row}"