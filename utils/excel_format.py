
def excel_output_format(results, baseline_inputs, WORKBOOK_FILE,TARGET_SHEET,TARGET_CELL,BASELINE_COMPARISON_CELL, original_forecast_value, base_target_value, wb_values):
    import pandas as pd
    import numpy as np


    print("\n📊 Top Attribution Results (% of total Δ):")

    top_results = sorted(results, key=lambda r: abs(r["delta"]), reverse=True)[:10]
    for r in top_results:
        print(f"{r['description']:<40} | % of Total: {r['normalized_pct']:+.2f}%")

    

    def format_number(value, sig_digits=2):
        """Format number to 2-3 significant digits"""
        if value is None or value == "":
            return ""
        try:
            num = float(value)
            if num == 0:
                return "0"
            # Round to significant digits
            from math import log10, floor
            rounded = round(num, -int(floor(log10(abs(num)))) + (sig_digits - 1))
            # Format nicely
            if abs(rounded) >= 1:
                return f"{rounded:.2f}".rstrip('0').rstrip('.')
            else:
                return f"{rounded:.3f}".rstrip('0').rstrip('.')
        except:
            return str(value)

    # Build a lookup dictionary from baseline_inputs for fast access
    cell_lookup = {}
    for bi in baseline_inputs:
        key = f"{bi['sheet']}!{bi['cell']}"
        cell_lookup[key] = {
            "original": format_number(bi.get("original_value")),
            "change_to": format_number(bi.get("change_to")),
            "formula": bi.get("formula", ""),
            "period": bi.get("period", ""),
            "info": bi.get("info", "")
        }

    # Define thresholds for meaningful effects
    MIN_DELTA_THRESHOLD = 1e-6  # Minimum absolute delta to be considered meaningful
    MIN_PCT_THRESHOLD = 0.01    # Minimum absolute percentage to be considered meaningful (0.01%)

    print(f"📊 Filtering results with thresholds:")
    print(f"   - Minimum |Δ Target|: {MIN_DELTA_THRESHOLD}")
    print(f"   - Minimum |% of Total|: {MIN_PCT_THRESHOLD}%")

    # Filter results to exclude zero/negligible effects
    meaningful_results = []
    excluded_count = 0
    excluded_total_delta = 0

    for r in results:
        abs_delta = abs(r['delta'])
        abs_pct = abs(r['normalized_pct'])
        
        # Keep if delta or percentage is above threshold
        if abs_delta >= MIN_DELTA_THRESHOLD and abs_pct >= MIN_PCT_THRESHOLD:
            meaningful_results.append(r)
        else:
            excluded_count += 1
            excluded_total_delta += r['delta']
            print(f"   ❌ Excluded: '{r['description'][:40]}...' (Δ={r['delta']:+.6f}, %={r['normalized_pct']:+.3f}%)")

    print(f"\n📈 Filtering Summary:")
    print(f"   ✅ Meaningful results: {len(meaningful_results)}")
    print(f"   ❌ Excluded (zero/negligible): {excluded_count}")
    print(f"   📉 Total excluded delta: {excluded_total_delta:+.6f}")

    # Add excluded effects as a summary line if there are any
    if excluded_count > 0 and abs(excluded_total_delta) >= MIN_DELTA_THRESHOLD:
        # Calculate percentage of total for excluded effects


        total_delta = original_forecast_value - base_target_value
        excluded_pct = (excluded_total_delta / total_delta * 100) if abs(total_delta) > 1e-6 else 0
        
        meaningful_results.append({
            "description": f"Minor effects (< {MIN_PCT_THRESHOLD}% each, {excluded_count} items)",
            "cells": [],
            "delta": excluded_total_delta,
            "normalized_pct": excluded_pct
        })
        print(f"   📝 Added summary line for minor effects: Δ={excluded_total_delta:+.6f}, %={excluded_pct:+.2f}%")

    # Create export data by combining attribution summary with baseline details
    export_data = []
    for r in meaningful_results:
        details = []
        for cell in r['cells']:
            meta = cell_lookup.get(cell, {})
            # Clean, spaced format
            details.append(
                f"{cell} | Period: {meta.get('period')} | "
                f"Original: {meta.get('original')} | "
                f"Change To: {meta.get('change_to')} | "
                f"Formula: {meta.get('formula')} | "
                f"Info: {meta.get('info')}"
            )

        # Parse numeric percent value for sorting
        numeric_pct = float(r['normalized_pct'])

        export_data.append({
            "Description": r["description"],
            "% of Total": f"{numeric_pct:+.2f}%",
            "Δ Target": f"{r['delta']:+.4f}",
            "Affected Cells": ", ".join(r["cells"]) if r["cells"] else "Multiple minor effects",
            "Details": "\n".join(details) if details else "Summary of multiple small effects",
            "_sort_key": abs(numeric_pct)
        })

    # Sort by absolute % contribution descending
    export_data_sorted = sorted(export_data, key=lambda x: -x["_sort_key"])

    # Remove helper key and export to Excel
    for row in export_data_sorted:
        del row["_sort_key"]

    df = pd.DataFrame(export_data_sorted)

    # Create filename with meaningful count
    meaningful_count = len([r for r in meaningful_results if r["cells"]])  # Exclude summary lines
    filename = f"meaningful_attribution_results_{meaningful_count}_effects.xlsx"

    # Export to Excel with formatting
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attribution Results', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Attribution Results']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with some padding
                adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Make header row bold
            from openpyxl.styles import Font
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
        
        print(f"\n✅ Exported {len(export_data_sorted)} meaningful results to {filename}")
        print(f"📊 Results include {meaningful_count} actual effects plus any summary lines")
        
    except Exception as e:
        # Fallback to basic export if formatting fails
        df.to_excel(filename, index=False)
        print(f"\n✅ Exported {len(export_data_sorted)} meaningful results to {filename} (basic format)")
        print(f"⚠️ Advanced formatting failed: {e}")


    # Show a preview of the top results
    print(f"\n🔝 Top 5 Meaningful Results:")
    for i, row in enumerate(export_data_sorted[:5], 1):
        print(f"  {i}. {row['Description'][:50]:<50} | {row['% of Total']:>8} | {row['Δ Target']:>10}")

    # Verification: Check that we haven't lost significant effects
    total_meaningful_delta = sum(r['delta'] for r in meaningful_results)
    total_original_delta = sum(r['delta'] for r in results)
    coverage = (total_meaningful_delta / total_original_delta * 100) if abs(total_original_delta) > 1e-6 else 100

    print(f"\n🎯 Verification:")
    print(f"   Original total effects: {len(results)}")
    print(f"   Meaningful effects kept: {len(meaningful_results)}")
    print(f"   Delta coverage: {coverage:.1f}% of original total")






    from openpyxl import Workbook 
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font
    import numpy as np
    import pandas as pd
    from openpyxl import load_workbook
    import os

    TOP_N = 6
    BLANK_ROWS_AFTER_TABLE = 5

    # === Format numbers ===
    def format_number_custom(x):
        try:
            x = float(x)
            if abs(x) >= 100:
                return f"{x:.0f}"
            else:
                return f"{x:.2f}"
        except:
            return x

    def format_pct(x):
        try:
            return f"{float(x):+.1f}%"
        except:
            return x

    # === Prepare export rows ===
    sorted_results = sorted(meaningful_results, key=lambda x: abs(x["normalized_pct"]), reverse=True)
    top_results = sorted_results[:TOP_N]
    grouped_results = sorted_results[TOP_N:]

    # Separate interrelated effects (set aside completely)
    interrelated = next((r for r in grouped_results if r["description"] == "Interrelated effects"), None)
    if interrelated:
        grouped_results = [r for r in grouped_results if r["description"] != "Interrelated effects"]

    # === Build export rows ===
    def build_export_row(r):
        details = []
        for cell in r.get("cells", []):
            meta = cell_lookup.get(cell, {})
            details.append(
                f"{cell} | Period: {meta.get('period')} | "
                f"Original: {meta.get('original')} | "
                f"Change To: {meta.get('change_to')} | "
                f"Formula: {meta.get('formula')} | "
                f"Info: {meta.get('info')}"
            )
        return {
            "Description": r["description"],
            "% of Total": format_pct(r["normalized_pct"]),
            "Δ Target": format_number_custom(r["delta"]),
            "Callable Link": "",
            "Affected Cells": ", ".join(r.get("cells", [])) if r.get("cells") else "Multiple minor effects",
            "Details": "\n".join(details) if details else "Summary of multiple small effects"
        }

    export_data = [build_export_row(r) for r in top_results]

    # Add grouped lower impact drivers (WITHOUT Interrelated Effects)
    if grouped_results:
        export_data.append({
            "Description": f"Lower Impact Drivers (n = {len(grouped_results)})",
            "% of Total": format_pct(sum(r['normalized_pct'] for r in grouped_results)),
            "Δ Target": format_number_custom(sum(r['delta'] for r in grouped_results)),
            "Callable Link": "",
            "Affected Cells": "",
            "Details": ""
            
        })
        export_data.extend(build_export_row(r) for r in grouped_results)

    # Add Interrelated Effects at the bottom (just before Sum row)
    if interrelated:
        export_data.append(build_export_row(interrelated))

    # Sum row (exclude Interrelated Effects)
    sum_delta = sum(r["delta"] for r in top_results + grouped_results)
    export_data.append({
        "Description": "Sum",
        "% of Total": "100.0%",
        "Δ Target": format_number_custom(sum_delta),
        "Callable Link": "",
        "Affected Cells": "",
        "Details": ""
    })

    # === Build Excel Workbook ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Attribution Results"
    #wb_values = load_workbook(WORKBOOK_FILE, data_only=True)

    # === Summary Block ===
    historical_actual = wb_values[TARGET_SHEET][BASELINE_COMPARISON_CELL].value
    current_forecasted = wb_values[TARGET_SHEET][TARGET_CELL].value  # or use `actual_2025` if already defined
    original_name = os.path.splitext(WORKBOOK_FILE)[0].split('/')[-1]


    #base_target_value = wb.sheets[TARGET_SHEET].range(TARGET_CELL).value


    ws.append([f"Result for {original_name}"])
    ws.append(["Historical Actual Value", format_number_custom(historical_actual)])
    ws.append(["Baseline Forecasted Value", format_number_custom(base_target_value)])
    ws.append(["Current Forecasted Value", format_number_custom(current_forecasted)])
    ws.append(["Change in Target", format_number_custom(original_forecast_value - base_target_value)])
    ws.append([])

    # === Export Table ===
    start_row = ws.max_row + 1
    df_final = pd.DataFrame(export_data)
    for row in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(row)

    # === Group lower impact rows (add outlining/collapsing functionality) ===
    group_header_row = None
    sum_row = None
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), start=start_row):
        val = row[0].value
        if val and isinstance(val, str):
            if val.startswith("Lower Impact Drivers"):
                group_header_row = i
            elif val == "Sum":
                sum_row = i

    if group_header_row and sum_row:
        # Find Interrelated Effects row (should be just before Sum)
        interrelated_row = None
        for i in range(sum_row - 1, group_header_row, -1):
            val = ws.cell(row=i, column=1).value
            if val and isinstance(val, str) and "Interrelated Effects" in val:
                interrelated_row = i
                break
        
        # Group only the Lower Impact Drivers detail rows (not Interrelated Effects)
        group_end_row = interrelated_row - 1 if interrelated_row else sum_row - 1
        
        for r in range(group_header_row + 1, group_end_row + 1):
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True
        ws.sheet_properties.outlinePr.summaryBelow = True

    # === Add blank rows ===
    for _ in range(BLANK_ROWS_AFTER_TABLE):
        ws.append([])

    # === EPS Impact Summary Block ===
    normalized_vals = [
        abs(float(r["% of Total"].strip('%')))
        for r in export_data if r["Description"] not in ("Sum",) and not r["Description"].startswith("Lower Impact")
    ]
    q75, q50, q25 = np.percentile(normalized_vals, [75, 50, 25]) if normalized_vals else (0, 0, 0)

    impact_summary = {
        "High Impact (Top 25%)": [],
        "Medium - High (Next 25%)": [],
        "Medium - Low (Next 25%)": [],
        "Low Impact (Bottom 25%)": []
    }
    for r in export_data:
        if r["Description"] in ("Sum",) or r["Description"].startswith("Lower Impact"):
            continue
        val = abs(float(r["% of Total"].strip('%')))
        if val >= q75:
            impact_summary["High Impact (Top 25%)"].append(r["Description"])
        elif val >= q50:
            impact_summary["Medium - High (Next 25%)"].append(r["Description"])
        elif val >= q25:
            impact_summary["Medium - Low (Next 25%)"].append(r["Description"])
        else:
            impact_summary["Low Impact (Bottom 25%)"].append(r["Description"])

    ws.append(["EPS Impact Summary"])
    for level, descriptions in impact_summary.items():
        ws.append([level, "Count", len(descriptions)])
        for desc in descriptions:
            ws.append(["", "", desc])
        ws.append([])

    # === Bold header row ===
    for cell in ws[f"A{start_row}":f"A{start_row}"][0]:
        cell.font = Font(bold=True)

    # === Column Widths ===
    column_widths = {
        "A": 72,
        "B": 12,
        "C": 12,
        "D": 8,
        "E": 20,
        "F": 24
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ####################### XLWINGS SECTION WITH VBA BUTTONS ##############################
    import xlwings as xw
    import os
    import shutil
    from datetime import datetime

    # Configuration
    TEMPLATE_PATH = "template_with_macro.xlsm"
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    FINAL_FILENAME = os.path.join(desktop_path, f"attribution_test_{original_name}.xlsm")
    print(desktop_path)
    print(FINAL_FILENAME)

    print(f"🚀 Starting VBA-enabled Excel file generation...")

    # Step 1: Save the openpyxl workbook as temporary xlsx file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_xlsx = f"temp_attribution_{timestamp}.xlsx"
    wb.save(temp_xlsx)
    print(f"💾 Saved temporary file: {temp_xlsx}")

    # Step 2: Copy template and prepare final file
    if os.path.exists(FINAL_FILENAME):
        try:
            os.remove(FINAL_FILENAME)
            print("🗑️ Removed existing output file")
        except PermissionError:
            print(f"❌ Cannot remove {FINAL_FILENAME} - file may be open")
            exit(1)

    print(TEMPLATE_PATH)
    print(FINAL_FILENAME)
    print(' ')


    shutil.copy(TEMPLATE_PATH, FINAL_FILENAME)
    print(f"✅ Copied macro template to: {FINAL_FILENAME}")

    # Step 3: Open both workbooks with xlwings
    try:
        app = xw.App(visible=False)  # Set to True for debugging
        
        # Open temporary xlsx file (source data)
        temp_book = app.books.open(os.path.abspath(temp_xlsx))
        temp_sheet = temp_book.sheets["Attribution Results"]
        
        # Open macro-enabled xlsm file (destination)
        final_book = app.books.open(os.path.abspath(FINAL_FILENAME))
        final_sheet = final_book.sheets["Attribution Results"]
        
        print("📖 Opened both workbooks successfully")
        
    except Exception as e:
        print(f"❌ Error opening workbooks: {e}")
        if 'app' in locals():
            app.quit()
        exit(1)

    try:
        # Step 4: Copy all data from temp workbook to final workbook
        used_range = temp_sheet.used_range
        if used_range is not None:
            # Copy all data at once
            data = used_range.value
            final_sheet.range("A1").value = data
            print("📊 Copied all data to macro-enabled workbook")
        
        # Step 5: Set original workbook filename in Z1
        final_sheet.range("Z1").value = WORKBOOK_FILE
        print(f"🏷️ Set original filename: {WORKBOOK_FILE}")
        
        # Step 6: Find the data table start row and sum row
        # Look for the main table headers (after EPS Summary)
        DATA_START_ROW = None
        sum_row = None
        
        # Read column A to find table structure
        col_a_range = final_sheet.range("A1:A50")  # Reasonable search range
        col_a_values = col_a_range.value
        
        for i, value in enumerate(col_a_values):
            if isinstance(value, str):
                if value == "Description":  # Table header
                    DATA_START_ROW = i + 1  # Excel is 1-indexed
                elif value.strip().lower() == "sum":
                    sum_row = i + 1
                    break
        
        if DATA_START_ROW is None:
            print("⚠️ Could not find table header 'Description'")
            DATA_START_ROW = 8  # Fallback
        
        if sum_row is None:
            print("⚠️ Could not find Sum row")
            sum_row = DATA_START_ROW + len(export_data)
        
        print(f"🎯 Table starts at row {DATA_START_ROW}, Sum at row {sum_row}")
        
        # Step 7: Read affected cells column and create buttons
        button_start_row = DATA_START_ROW + 1  # Skip header
        button_end_row = sum_row - 1  # Don't include Sum row
        
        if button_end_row >= button_start_row:
            # Read affected cells column (Column E) in the range
            affected_cells_range = final_sheet.range(f"E{button_start_row}:E{button_end_row}")
            affected_cells_values = affected_cells_range.value
            
            # Ensure it's a list even if single row
            if not isinstance(affected_cells_values, list):
                affected_cells_values = [affected_cells_values]
            
            buttons_created = 0
            
            # Create buttons for valid rows
            for i, affected_cells in enumerate(affected_cells_values):
                current_row = button_start_row + i
                
                # Skip if no affected cells or invalid format
                if not affected_cells or not isinstance(affected_cells, str):
                    print(f"  ⏭️ Row {current_row}: No affected cells")
                    continue
                    
                if "!" not in affected_cells:
                    print(f"  ⏭️ Row {current_row}: No sheet reference")
                    continue
                    
                if "multiple minor effects" in affected_cells.lower():
                    print(f"  ⏭️ Row {current_row}: Multiple minor effects")
                    continue
                
                # Check if button already exists
                button_name = f"btn_{current_row}"
                button_exists = False
                try:
                    existing_button = final_sheet.api.Buttons(button_name)
                    button_exists = True
                    print(f"  ♻️ Row {current_row}: Button already exists")
                except:
                    button_exists = False
                
                if not button_exists:
                    # ✅ Create the button
                    try:
                        target_cell = final_sheet.range(f"D{current_row}")
                        left = target_cell.left
                        top = target_cell.top
                        width = 45
                        height = 15

                        # Extract button label: first cell reference without sheet
                        # E.g., "Sheet1!B2, Sheet2!C4" → "B2"
                        first_ref = affected_cells.split(',')[0].strip()  # Take first entry
                        if '!' in first_ref:
                            first_ref = first_ref.split('!')[1]  # Remove sheet name

                        button = final_sheet.api.Buttons().Add(left, top, width, height)
                        button.Text = first_ref  # Set the extracted reference as label
                        button.Name = button_name
                        button.OnAction = "GoToAffectedCell"

                        buttons_created += 1

                        print(f"  ✅ Row {current_row}: Created button for '{affected_cells[:50]}...'")
                        
                    except Exception as e:
                        print(f"  ❌ Row {current_row}: Error creating button - {e}")
            
            print(f"🎯 Created {buttons_created} buttons total")
        
        # Step 8: Apply grouping/outlining to Lower Impact Drivers
        print("🔍 Applying grouping to Lower Impact Drivers...")
        
        # Find rows for grouping in the xlwings workbook
        group_header_row = None
        interrelated_row = None
        sum_row_xlwings = None
        
        # Read column A to find structure
        col_a_range = final_sheet.range("A1:A50")
        col_a_values = col_a_range.value
        
        for i, value in enumerate(col_a_values):
            if isinstance(value, str):
                excel_row = i + 1  # Excel is 1-indexed
                print(f"  Row {excel_row}: '{value}'")
                
                if value.startswith("Lower Impact Drivers"):
                    group_header_row = excel_row
                    print(f"    ✅ Found group header at row {excel_row}")
                elif "Interrelated effects" in value:
                    interrelated_row = excel_row
                    print(f"    ✅ Found interrelated at row {excel_row}")
                elif value == "Sum":
                    sum_row_xlwings = excel_row
                    print(f"    ✅ Found sum at row {excel_row}")
        
        # Apply grouping if we found the necessary rows
        if group_header_row and sum_row_xlwings:
            group_start_row = group_header_row + 1
            group_end_row = interrelated_row - 1 if interrelated_row else sum_row_xlwings - 1
            
            print(f"🎯 Grouping rows {group_start_row} to {group_end_row}")
            
            if group_start_row <= group_end_row:
                try:
                    # Select the range to group
                    group_range = final_sheet.range(f"{group_start_row}:{group_end_row}")
                    
                    # Apply grouping
                    group_range.api.Group()
                    print(f"✅ Successfully grouped rows {group_start_row}-{group_end_row}")
                    
                    # Collapse the group (this is the correct way)
                    try:
                        # Set outline level to collapsed state
                        final_sheet.api.Outline.ShowLevels(RowLevels=1)
                        print("✅ Successfully collapsed the group")
                    except Exception as collapse_error:
                        print(f"⚠️ Group created but couldn't collapse: {collapse_error}")
                    
                except Exception as e:
                    print(f"❌ Error applying grouping: {e}")
            else:
                print("⚠️ No valid rows to group")
        else:
            print("⚠️ Could not find required rows for grouping")
            print(f"   group_header_row: {group_header_row}")
            print(f"   sum_row_xlwings: {sum_row_xlwings}")
        
        # Step 9: Set column widths in the xlwings workbook
        column_widths = {
            "A": 72,
            "B": 12, 
            "C": 12,
            "D": 9,
            "E": 32,
            "F": 24
        }
        
        for col_letter, width in column_widths.items():
            final_sheet.range(f"{col_letter}:{col_letter}").column_width = width
        
        print("📏 Applied column widths to final workbook")
        
        # Step 10: Save and close
        final_book.save()
        print("💾 Saved macro-enabled workbook")
        
    except Exception as e:
        print(f"❌ Error during processing: {e}")
        import traceback
        traceback.print_exc()

    finally:
        # Always close Excel
        try:
            temp_book.close()
            final_book.close()
            app.quit()
            print("🔒 Closed Excel application")
        except:
            pass
        
        # Clean up temporary file
        try:
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
                print("🗑️ Cleaned up temporary file")
        except:
            pass

    print(f"\n✅ COMPLETED: {FINAL_FILENAME}")
    print("📋 Usage Instructions:")
    print("1. Ensure your original Excel file is open")
    print(f"2. Open {FINAL_FILENAME}")
    print("3. Click any 'Button' in the Callable Link column")
    print("4. VBA will jump to the original file and highlight cells")
    print(f"5. Original attribution count: {meaningful_count} effects")