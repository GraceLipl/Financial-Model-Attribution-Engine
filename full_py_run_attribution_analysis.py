import sys
import os
import time
import re
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries



def extract_numeric_constants(formula):
    if not formula or not formula.startswith("="):
        return []

    # Remove quoted strings (e.g., text literals in formulas)
    cleaned = re.sub(r'"[^"]*"', '', formula)

    # Tokenize the formula into pieces split by operators, parentheses, commas, spaces
    tokens = re.split(r'[\+\-\*/\^\(\),\s]+', cleaned)

    constants = []
    for token in tokens:
        token = token.strip()
        # Skip if it's a cell reference like A1, AA123, etc.
        if re.match(r'^[A-Za-z]{1,3}[0-9]{1,5}$', token):
            continue
        try:
            constants.append(float(token))
        except ValueError:
            pass  # ignore non-numeric tokens
    return constants


def resolve_label_with_hierarchy(wb_values, sheet_name, row):
    """
    Get cell description from columns A, B, C without bold header enrichment
    Modified to work with wb_values parameter and sheet_name
    """
    print(f"DEBUG: Getting description for row {row}")
    
    # Get the sheet object from wb_values
    sheet = wb_values[sheet_name]
    
    # Check columns A, B, C in order to find the first non-empty cell
    for col_letter in ['A', 'B', 'C']:
        cell = sheet[f"{col_letter}{row}"]
        print(f"DEBUG: Checking {col_letter}{row} = {cell.value}")
        
        if cell.value is not None:
            cell_value = cell.value
            
            # If cell has a formula, follow it
            if hasattr(cell, 'data_type') and (cell.data_type == "f" or (isinstance(cell_value, str) and cell_value.startswith("="))):
                print(f"DEBUG: Found formula in {col_letter}{row}: {cell_value}")
                m = re.match(r"=([^!]+)!([A-Z]+[0-9]+)", str(cell_value))
                if m:
                    target_sheet, target_cell = m.groups()
                    try:
                        # Use the wb_values parameter instead of global wb_values
                        cell_value = wb_values[target_sheet][target_cell].value
                        print(f"DEBUG: Formula resolved to: {cell_value}")
                        if cell_value is None:
                            cell_value = ""
                        else:
                            cell_value = str(cell_value).strip()
                    except:
                        cell_value = ""
            
            result = str(cell_value).strip() if cell_value else ""
            print(f"DEBUG: Final result for row {row}: '{result}'")
            return result
    
    # If no cell found, return row number
    print(f"DEBUG: No cell found for row {row}, returning Row {row}")
    return f"Row {row}"



def build_column_period_map(wb_values, sheet_name, start_row=1, today_date=None, max_row=150):
    """
    Build column to period map (automated) - EXACTLY as in your Jupyter notebook
    """
    if today_date is None:
        today_date = datetime.date.today()

    sheet = wb_values[sheet_name]  # Access worksheet by name from Workbook
    col_to_period = {}

    # Enhanced pattern to match more period formats - same as lookup_column_period
    pattern = re.compile(r'\b(\d{1,2}Q\d{2,4}E?|\d{4}E?|FY\d{2,4}E?)\b', re.I)

    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        period = ""

        for row in range(start_row, min(max_row + 1, sheet.max_row + 1)):
            cell_obj = sheet[f"{col_letter}{row}"]
            cell_value = cell_obj.value

            if cell_value is None:
                continue

            # Process string values - look for period patterns
            if isinstance(cell_value, str):
                text = cell_value.strip()
                
                # Search for period pattern in the text
                match = pattern.search(text)
                if match:
                    period = match.group(1)  # Return the matched period
                    break

            # Process numeric values - handle years stored as numbers
            elif isinstance(cell_value, (int, float)):
                # Check if it's a 4-digit year
                if 2000 <= cell_value <= 2050:
                    year = int(cell_value)
                    # Determine if it's forecast based on year threshold
                    forecast_threshold = 2025  # Years >= 2025 are considered forecast
                    if year >= forecast_threshold:
                        period = f"{year}E"
                    else:
                        period = str(year)
                    break

            # Skip other types (dates, etc.)

        if period:
            col_to_period[col_letter] = period

    return col_to_period

def build_all_sheets_col_to_period(wb_values, today_date=None, max_row=150):
    """
    Build col_to_period maps for all sheets in the workbook - EXACTLY as in your Jupyter notebook
    """
    all_col_to_period = {}
    
    for sheet_name in wb_values.sheetnames:  # Use .sheetnames for openpyxl Workbook
        try:
            col_to_period = build_column_period_map(
                wb_values, sheet_name, start_row=1, today_date=today_date, max_row=max_row
            )
            if col_to_period:  # Only store if we found periods
                all_col_to_period[sheet_name] = col_to_period
                print(f"✓ {sheet_name}: Found {len(col_to_period)} period columns")
            else:
                print(f"⚠ {sheet_name}: No periods found")
        except Exception as e:
            print(f"✗ {sheet_name}: Error - {e}")
    
    return all_col_to_period

def lookup_column_period_from_map(all_col_to_period, sheet_name, cell_ref):
    """
    Helper function to get period for a specific cell from the all_col_to_period map
    """
    if sheet_name not in all_col_to_period:
        return ""
    
    col_letter = re.match(r"[A-Z]+", cell_ref).group(0)
    return all_col_to_period[sheet_name].get(col_letter, "")


import re
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils import range_boundaries

def parse_sheet_reference(ref):
    """
    Parse a sheet reference like 'Sheet Name'!A1 or Sheet!A1
    Returns (sheet_name, cell_ref)
    """
    if "!" not in ref:
        return None, ref
    
    # Split by the last occurrence of ! to handle cases like 'Sheet!Name'!A1
    parts = ref.rsplit("!", 1)
    sheet_part = parts[0]
    cell_part = parts[1]
    
    # Remove quotes if present
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1]
    
    return sheet_part, cell_part




def trace_merged(wb_values, wb_formulas, sheet_name, cell_ref, all_col_to_period, max_depth=20, visited=None, depth=0, path=None):
    def is_terminal_cell(sheet_name, cell_ref, formula):
        col_letter = re.match(r"[A-Z]+", cell_ref).group(0)
        # Updated to use the new lookup function
        period = lookup_column_period_from_map(all_col_to_period, sheet_name, cell_ref)
        if not period:
            return True
        if sheet_name.lower().startswith("schedule"):
            return True
        if formula is None:
            return True
        if isinstance(formula, str) and "[" in formula:
            return True
        if re.match(r"^=[0-9.\-+*/ ()]+$", formula):
            return True
        return False

    if visited is None:
        visited = set()
    if path is None:
        path = []

    key = (sheet_name, cell_ref)
    if key in visited or depth > max_depth:
        return []

    visited.add(key)
    current_path = path + [f"{sheet_name}!{cell_ref}"]

    sheet_v = wb_values[sheet_name]
    sheet_f = wb_formulas[sheet_name]
    cell_v = sheet_v[cell_ref]

    try:
        cell_f = sheet_f[cell_ref]
        formula = cell_f.value if isinstance(cell_f.value, str) and cell_f.value.startswith("=") else None
        description = resolve_label_with_hierarchy(wb_values, sheet_name, cell_f.row) or f"Row {cell_f.row}"
    except KeyError:
        formula = None
        description = None

    value = cell_v.value
    external_ref_segment = ""
    is_external = False

    if formula and "[" in formula:
        match_ref = re.search(r"(\[[^\]]+\][^\+\-\*/\)^\n]*)", formula)
        external_ref_segment = match_ref.group(1) if match_ref else "[EXTERNAL]"
        is_external = True

    is_terminal = is_terminal_cell(sheet_name, cell_ref, formula)

    numeric_literals = []
    if not is_external and formula:
        # Remove quoted strings first
        cleaned_formula = re.sub(r'"[^"]*"', '', formula)
        
        # Find all cell references (including those with $ and sheet names)
        cell_refs = re.findall(r'(?<![0-9A-Za-z_])(?:\'[^\']*\'!\$?[A-Z]{1,3}\$?[0-9]{1,5}|[A-Za-z0-9_]+!\$?[A-Z]{1,3}\$?[0-9]{1,5}|\$?[A-Z]{1,3}\$?[0-9]{1,5})(?![0-9A-Za-z_])', cleaned_formula)
        
        # Remove all cell references from the formula to isolate numeric constants
        temp_formula = cleaned_formula
        for ref in cell_refs:
            temp_formula = temp_formula.replace(ref, '')
        
        # Extract numeric constants
        all_numbers = re.findall(r'[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?', temp_formula)
        filtered_constants = []
        for num_str in all_numbers:
            try:
                val = float(num_str)
                if val == 1.0 and (re.search(r'\(1\s*\+\s*[\w\d]', formula) or re.search(r'[\w\d]\s*\+\s*1\)', formula)):
                    continue
                if val == -1.0 and formula.strip().endswith("-1"):
                    continue
                filtered_constants.append(val)
            except:
                continue
        numeric_literals = filtered_constants

    # Updated period lookup - use the new function
    period = lookup_column_period_from_map(all_col_to_period, sheet_name, cell_ref)

    output = [{
        "sheet": sheet_name,
        "cell": cell_ref,
        "formula": formula,
        "value": value,
        "depth": depth,
        "description": description,
        "is_terminal": is_terminal or is_external,
        "external_ref_segment": external_ref_segment,
        "constants": numeric_literals,
        "path": current_path,
        "period": period
    }]

    if is_terminal or is_external:
        return output

    refs = []
    ranges = []
    if isinstance(formula, str):
        # Updated regex to handle quoted sheet names AND $ symbols for absolute references
        # Pattern explanation:
        # - (?:'[^']*'!\$?[A-Z]{1,3}\$?[0-9]{1,5}) - quoted sheet with absolute refs
        # - (?:[A-Za-z0-9_]+!\$?[A-Z]{1,3}\$?[0-9]{1,5}) - unquoted sheet with absolute refs  
        # - (?:\$?[A-Z]{1,3}\$?[0-9]{1,5}) - local cell with absolute refs
        refs = re.findall(r"(?<![0-9A-Za-z_])(?:'[^']*'!\$?[A-Z]{1,3}\$?[0-9]{1,5}|[A-Za-z0-9_]+!\$?[A-Z]{1,3}\$?[0-9]{1,5}|\$?[A-Z]{1,3}\$?[0-9]{1,5})(?![0-9A-Za-z_])", formula)
        refs = [ref.replace("$", "") for ref in refs]
        
        # Updated regex for ranges with quoted sheet names AND $ symbols
        ranges = re.findall(r"('[^']*'!\$?[A-Z]{1,3}\$?[0-9]{1,5}:\$?[A-Z]{1,3}\$?[0-9]{1,5}|[A-Za-z0-9_]+!\$?[A-Z]{1,3}\$?[0-9]{1,5}:\$?[A-Z]{1,3}\$?[0-9]{1,5}|\$?[A-Z]{1,3}\$?[0-9]{1,5}:\$?[A-Z]{1,3}\$?[0-9]{1,5})", formula)
        ranges = [rng.replace("$", "") for rng in ranges]

    for ref in refs:
        # Use the new parsing function
        other_sheet, other_cell = parse_sheet_reference(ref)
        if other_sheet is None:
            other_sheet = sheet_name
            other_cell = ref

        col_letter = re.match(r"[A-Z]+", other_cell).group(0)

        # Updated filtering logic
        if other_sheet.lower().startswith("schedule"):
            continue
        
        # Check if this sheet has any period mapping at all
        if other_sheet not in all_col_to_period:
            continue
            
        # Check if this specific column has a period
        if col_letter not in all_col_to_period[other_sheet]:
            continue

        # Check if the sheet exists before trying to trace
        if other_sheet not in wb_values:
            print(f"Warning: Sheet '{other_sheet}' not found in workbook")
            continue

        output += trace_merged(wb_values, wb_formulas, other_sheet, other_cell, all_col_to_period, max_depth, visited, depth + 1, current_path)

    for rng in ranges:
        # Use the new parsing function for ranges too
        other_sheet, cell_range = parse_sheet_reference(rng)
        if other_sheet is None:
            other_sheet = sheet_name
            cell_range = rng

        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                col_letter = get_column_letter(col)
                
                # Updated filtering logic for ranges
                if other_sheet.lower().startswith("schedule"):
                    continue
                
                # Check if this sheet has any period mapping at all
                if other_sheet not in all_col_to_period:
                    continue
                    
                # Check if this specific column has a period
                if col_letter not in all_col_to_period[other_sheet]:
                    continue
                
                # Check if the sheet exists before trying to trace
                if other_sheet not in wb_values:
                    print(f"Warning: Sheet '{other_sheet}' not found in workbook")
                    continue
                    
                cell = f"{col_letter}{row}"
                output += trace_merged(wb_values, wb_formulas, other_sheet, cell, all_col_to_period, max_depth, visited, depth + 1, current_path)

    return output



def is_effectively_zero(value, tolerance=1e-10):
    """Check if a value is effectively zero, accounting for floating-point precision"""
    if value is None:
        return True
    if isinstance(value, (int, float)):
        return abs(value) < tolerance
    return False

def is_error_value(value):
    """Check if a value is an Excel error (starts with #)"""
    if value is None:
        return False
    return str(value).startswith("#")



def is_terminal_input(item, trace, all_col_to_period):
    period = item["period"]
    if not period.endswith("E"):
        return False

    formula = item.get("formula")
    constants = item.get("constants", [])
    
    excluded_constants = {0.25, 0.33, 0.5, 0.67, 0.75, 1/4, 1/3, 1/2, 2/3, 3/4, 
                          1, -1, 2, 3, 4, 52, 90, 100, 
                          180, 270, 365, 1000, 10000, 100000, 1000000}
    meaningful_constants = [c for c in constants if c not in excluded_constants]
    has_meaningful_constants = bool(meaningful_constants)

    if not formula or "[" in formula:
        return True

    if re.match(r"^=[0-9.\-+*/ ()]+$", formula):
        return True

    if has_meaningful_constants:
        complex_keywords = ["IF", "CONCATENATE", "VLOOKUP", "HLOOKUP", "INDEX", "MATCH", 
                            "CHOOSE", "OFFSET", "INDIRECT", "SUMIF", "COUNTIF", "AVERAGEIF"]
        formula_upper = formula.upper()
        if any(keyword in formula_upper for keyword in complex_keywords):
            return False
        return True

    refs = re.findall(
        r'(?<![0-9A-Za-z_])(?:\'[^\']*\'!\$?[A-Z]{1,3}\$?[0-9]{1,5}|[A-Za-z0-9_]+!\$?[A-Z]{1,3}\$?[0-9]{1,5}|\$?[A-Z]{1,3}\$?[0-9]{1,5})(?![0-9A-Za-z_])',
        formula
    )

    for ref in refs:
        ref = ref.replace("$", "")
        ref_sheet, ref_cell = parse_sheet_reference(ref)
        if ref_sheet is None:
            ref_sheet = item['sheet']

        ref_found = False
        for ref_item in trace:
            if ref_item["sheet"] == ref_sheet and ref_item["cell"] == ref_cell:
                ref_found = True
                if ref_item["period"].endswith("E"):
                    return False
                break

        if not ref_found:
            ref_period = lookup_column_period_from_map(all_col_to_period, ref_sheet, ref_cell)
            if ref_period.endswith("E"):
                return False

    return True



import re
from collections import defaultdict

def get_terminal_cell_path(terminal_item):
    return f"{terminal_item['sheet']}!{terminal_item['cell']}"

def find_terminal_chain(filtered_trace, terminal_item):
    """
    Find the full path (chain of cell references) for a given terminal item.
    """
    terminal_path = get_terminal_cell_path(terminal_item)
    for item in filtered_trace:
        if get_terminal_cell_path(item) == terminal_path:
            return item["path"]
    return []

def filter_chain_by_period(chain_paths, terminal_item, all_col_to_period):
    """
    Keep only items in the same period as the terminal.
    Returns list of (sheet, cell) tuples AND the filtered path strings for printing.
    """
    terminal_period = terminal_item["period"]
    filtered_tuples = []
    filtered_paths = []
    
    for path in chain_paths:
        sheet, cell = path.split("!")
        period = lookup_column_period_from_map(all_col_to_period, sheet, cell)
        if period == terminal_period:
            filtered_tuples.append((sheet, cell))
            filtered_paths.append(path)
    
    return filtered_tuples, filtered_paths

def get_trace_lookup(trace):
    return {f"{item['sheet']}!{item['cell']}": item for item in trace}

def get_nearest_parents_descriptions(filtered_chain, terminal_item, trace_lookup, num_parents=2):
    """
    From filtered chain, return up to two parent descriptions above the terminal cell.
    Also tries to find adjacent row descriptions if chain is too short.
    """
    terminal_path = get_terminal_cell_path(terminal_item)
    terminal_row = int(re.match(r"[A-Z]+(\d+)", terminal_item["cell"]).group(1))
    terminal_sheet = terminal_item["sheet"]
    terminal_col = re.match(r"([A-Z]+)\d+", terminal_item["cell"]).group(1)

    descriptions = []
    
    # First, try to get descriptions from the filtered chain
    for sheet, cell in reversed(filtered_chain):
        path = f"{sheet}!{cell}"
        if path == terminal_path:
            continue
        row = int(re.match(r"[A-Z]+(\d+)", cell).group(1))
        if sheet == terminal_sheet and row == terminal_row:
            continue
        desc = trace_lookup.get(path, {}).get("description", "")
        if desc and desc not in descriptions:
            descriptions.append(desc)
        if len(descriptions) == num_parents:
            break
    
    # If we don't have enough descriptions from chain, try adjacent rows
    if len(descriptions) < num_parents and len(filtered_chain) <= 1:
        print(f"DEBUG PARENTS: Chain too short ({len(filtered_chain)}), checking adjacent rows for {terminal_path}")
        
        # Check rows above the terminal (row-1, row-2, etc.)
        for offset in range(1, 5):  # Check up to 4 rows above
            if len(descriptions) >= num_parents:
                break
            adjacent_row = terminal_row - offset
            if adjacent_row > 0:
                adjacent_path = f"{terminal_sheet}!{terminal_col}{adjacent_row}"
                desc = trace_lookup.get(adjacent_path, {}).get("description", "")
                if desc and desc not in descriptions and desc != terminal_item.get("description", ""):
                    descriptions.append(desc)
                    print(f"DEBUG PARENTS: Found adjacent row description: {adjacent_path} -> '{desc}'")
    
    return list(reversed(descriptions))  # Return in order: parent1, parent2

def get_enriched_terminal_description(terminal_item, trace, all_col_to_period):
    """
    Enhanced version that returns both enriched description and chain info for printing.
    Format: branch → subbranch → terminal_description
    """
    trace_lookup = get_trace_lookup(trace)
    terminal_path = get_terminal_cell_path(terminal_item)

    full_chain = find_terminal_chain(trace, terminal_item)
    if not full_chain:
        return terminal_item.get("description", ""), None, None

    # Get both filtered tuples and paths for printing
    filtered_chain, filtered_paths = filter_chain_by_period(full_chain, terminal_item, all_col_to_period)
    
    if not filtered_chain:
        return terminal_item.get("description", ""), None, None

    parents = get_nearest_parents_descriptions(filtered_chain, terminal_item, trace_lookup)
    
    # Build enriched description: parents + terminal description (terminal at the end)
    terminal_desc = terminal_item.get("description", "")
    enriched_parts = parents + [terminal_desc] if terminal_desc else parents
    enriched_desc = " → ".join(p for p in enriched_parts if p)
    
    # If no enrichment occurred (no parents found), just return original description
    if not parents:
        enriched_desc = terminal_desc
    
    # Return enriched description, full chain info, and filtered chain info
    chain_info = {
        "full_chain": full_chain,
        "filtered_chain": filtered_paths,
        "full_length": len(full_chain),
        "filtered_length": len(filtered_paths)
    }
    
    return enriched_desc, chain_info, parents




import datetime
from openpyxl.utils import column_index_from_string, get_column_letter

def should_force_zero(value, description, number_format):
    """
    Returns True if the value and description indicate a growth/change input that should be forced to 0.
    """
    if not isinstance(value, (int, float)):
        return False

    if not description:
        return False

    desc_lower = description.lower()

    # Keywords for growth/change situationsa
    keywords = [
        "% change", "growth", "yoy", "yr/yr", "q/q", "mom", "y/y", "m/m",
        "variance", "var", "delta", "change", "chg", "chng",
        "inflation", "cpi", "ppi", "fx impact", "discount rate change",
        "wacc delta", "wacc change", "risk premium change",
        "margin delta", "cost increase rate", "tax change", "tax rate change"
    ]

    for kw in keywords:
        if kw in desc_lower:
            if "margin %" in desc_lower or "tax rate %" in desc_lower:
                return False
            return True

    return False


def parse_sheet_reference(ref):
    """
    Parse a sheet reference like 'Sheet Name'!A1 or Sheet!A1
    Returns (sheet_name, cell_ref)
    """
    if "!" not in ref:
        return None, ref
    
    # Split by the last occurrence of ! to handle cases like 'Sheet!Name'!A1
    parts = ref.rsplit("!", 1)
    sheet_part = parts[0]
    cell_part = parts[1]
    
    # Remove quotes if present
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1]
    
    return sheet_part, cell_part



def run_attribution_analysis_complete(baseline_inputs, workbook_path, target_sheet, target_cell):
    """
    Complete attribution analysis: No-copy connection logic + Full Jupyter attribution logic
    """
    import xlwings as xw
    import time
    import os
    import re
    from datetime import datetime
    from collections import defaultdict
    import traceback

    # Store all output for return
    attribution_output = []
    
    def capture_print(msg):
        """Capture print output for return to main function"""
        msg_str = str(msg)
        print(msg_str)  # Still print to console
        attribution_output.append(msg_str)

    # === Filter invalid baseline inputs (from Jupyter) ===
    baseline_inputs = [item for item in baseline_inputs if item["change_to"] is not None]

    capture_print("🚀 Starting complete attribution analysis...")
    capture_print("⚠️ Working directly on original file - will restore all changes at end")
    capture_print(f"📋 Processing {len(baseline_inputs)} baseline inputs")
    capture_print(f"🎯 Target: {target_sheet}!{target_cell}")
    capture_print(f"📂 Original file: {workbook_path}")

    # === STEP 1-2: Connect to existing Excel workbook (from working no-copy version) ===
    wb = None
    app = None
    original_values_backup = {}  # Store original values for restoration
    
    try:
        capture_print("📂 Connecting to Excel workbook...")
        
        # Strategy: Connect to already open workbook
        try:
            # First, try to find existing Excel apps
            apps = xw.apps
            if len(apps) > 0:
                app = apps[0]
                capture_print("   Found existing Excel application")
                
                # Look for the workbook by name
                workbook_name = os.path.basename(workbook_path)
                wb = None
                
                for book in app.books:
                    if book.name == workbook_name or book.fullname == workbook_path:
                        wb = book
                        capture_print(f"   Found open workbook: {book.name}")
                        break
                
                if wb is None:
                    # Open the file in existing Excel
                    capture_print(f"   Opening {workbook_name} in existing Excel...")
                    wb = app.books.open(workbook_path, update_links=False)
            else:
                raise Exception("No Excel apps found")
                
        except Exception as e:
            # Fallback: Create new Excel and open file
            capture_print(f"   Creating new Excel application... ({str(e)})")
            app = xw.App(visible=True, add_book=False)
            wb = app.books.open(workbook_path, update_links=False)
        
        if wb is None:
            raise Exception("Could not open workbook")
            
        capture_print("✅ Excel workbook connected successfully")
        
        # Configure Excel settings (non-critical)
        try:
            app.screen_updating = False
            app.enable_events = False
            app.display_alerts = False
            app.calculation = 'manual'
            capture_print("   Excel settings configured")
        except:
            capture_print("   ⚠️ Some Excel settings could not be configured")

        # === STEP 3: Get original forecast value (from working no-copy version) ===
        capture_print("📌 Reading original forecast value...")
        original_forecast_value = wb.sheets[target_sheet].range(target_cell).value
        
        if original_forecast_value is None:
            raise Exception(f"Target cell {target_sheet}!{target_cell} is empty or not accessible")
            
        capture_print(f"📌 Original Forecast Target Value: {original_forecast_value:.4f}")

        # === STEP 4: Backup original values AND FORMULAS ===
        capture_print("💾 Backing up original values and formulas...")
        
        for item in baseline_inputs:
            sheet_name = item["sheet"]
            cell_ref = item["cell"]
            
            try:
                rng = wb.sheets[sheet_name].range(cell_ref)
                original_val = rng.value  # Get the calculated value
                original_formula = rng.formula  # Get the formula (this was missing!)
                
                original_values_backup[f"{sheet_name}!{cell_ref}"] = {
                    "value": original_val,
                    "formula": original_formula  # Now backing up formulas too!
                }
                
                # Debug: Show what we're backing up for first few cells
                if len(original_values_backup) < 5:
                    if original_formula and original_formula.startswith('='):
                        capture_print(f"   Backing up {sheet_name}!{cell_ref}: formula='{original_formula}', value={original_val}")
                    else:
                        capture_print(f"   Backing up {sheet_name}!{cell_ref}: no formula, value={original_val}")
                        
            except Exception as e:
                capture_print(f"⚠️ Could not backup {sheet_name}!{cell_ref}: {e}")

        capture_print(f"💾 Backed up {len(original_values_backup)} cell values and formulas")

        # === STEP 5: Apply baseline changes (from Jupyter logic) ===
        capture_print("🔧 Applying baseline changes...")
        changes_applied = 0
        changes_verified = 0

        for i, item in enumerate(baseline_inputs):
            sheet, cell, change_to = item["sheet"], item["cell"], item["change_to"]
            original_value = item["original_value"]
            
            try:
                rng = wb.sheets[sheet].range(cell)
                
                # Check current value before change
                current_value = rng.value
                
                # Apply change
                rng.formula = None
                rng.value = change_to
                changes_applied += 1
                
                # Verify the change was applied
                new_value = rng.value
                if abs(new_value - change_to) < 1e-10:
                    changes_verified += 1
                
                # Show first few changes for debugging
                if i < 5:
                    capture_print(f"   {sheet}!{cell}: {original_value:.6f} → {current_value:.6f} → {new_value:.6f} (target: {change_to:.6f})")
                elif i == 5:
                    capture_print("   ...")
                
                if i % 10 == 0:  # Progress indicator
                    capture_print(f"   Applied {i+1}/{len(baseline_inputs)} changes...")
                    
            except Exception as e:
                capture_print(f"⚠️ Error applying baseline to {sheet}!{cell}: {e}")

        capture_print(f"📊 Changes applied: {changes_applied}/{len(baseline_inputs)}")
        capture_print(f"📊 Changes verified: {changes_verified}/{len(baseline_inputs)}")

        # === STEP 6: Force calculation (from Jupyter logic with no-copy robustness) ===
        capture_print("\n🔄 Forcing calculation after all baseline changes...")
        try:
            app.calculate()
            time.sleep(1)
            # Force a second calculation to ensure all dependencies are updated
            app.calculate()
            time.sleep(1)
            capture_print("✅ Forced calculations completed")
        except Exception as e:
            capture_print(f"⚠️ Error during forced calculation: {e}")

        # === STEP 7: Calculate baseline (from Jupyter logic) ===
        capture_print("\n🧮 Calculating baseline...")
        for i in range(5):  # Increased attempts
            try:
                capture_print(f"   Calculation attempt {i+1}...")
                app.calculate()
                time.sleep(2)
                
                # Check if target value changed
                current_target = wb.sheets[target_sheet].range(target_cell).value
                if current_target is not None:
                    capture_print(f"   Target after calc {i+1}: {current_target:.6f}")
                    
                    if abs(current_target - original_forecast_value) > 1e-10:
                        capture_print("   ✅ Target value changed, calculation successful")
                        break
                    else:
                        capture_print("   ⚠️ Target unchanged, trying again...")
                else:
                    capture_print(f"   ⚠️ Target is None after calculation {i+1}")
                    
            except Exception as e:
                capture_print(f"⚠️ Excel busy (attempt {i+1}): {e}")
                time.sleep(3)

        base_target_value = wb.sheets[target_sheet].range(target_cell).value
        if base_target_value is None:
            capture_print("❌ Cannot proceed - baseline target value is None")
            raise Exception("Baseline target value is None after calculations")
            
        capture_print(f"📌 Baseline Target Value: {base_target_value:.6f}")

        # === STEP 8: Diagnostic checks (from Jupyter logic) ===
        capture_print("\n🔍 Diagnostic: Verifying baseline changes are still applied...")
        still_applied = 0
        reverted_count = 0

        for i, item in enumerate(baseline_inputs[:10]):  # Check first 10 for speed
            sheet, cell, change_to = item["sheet"], item["cell"], item["change_to"]
            original_value = item["original_value"]
            
            try:
                current_value = wb.sheets[sheet].range(cell).value
                if current_value is not None:
                    if abs(current_value - change_to) < 1e-10:
                        still_applied += 1
                    elif abs(current_value - original_value) < 1e-10:
                        reverted_count += 1
                    
                    if i < 5:
                        capture_print(f"   {sheet}!{cell}: current={current_value:.6f}, should_be={change_to:.6f}, original={original_value:.6f}")
            except Exception as e:
                capture_print(f"⚠️ Error checking {sheet}!{cell}: {e}")

        capture_print(f"📊 First 10 cells: {still_applied} still changed, {reverted_count} reverted to original")

        # === STEP 9: Check calculation mode (from Jupyter logic) ===
        capture_print(f"\n🔍 Excel calculation mode: {app.calculation}")
        if app.calculation != 'manual':
            capture_print("⚠️ Calculation mode is not manual, setting to manual...")
            try:
                app.calculation = 'manual'
                capture_print("✅ Set to manual calculation")
            except Exception as e:
                capture_print(f"⚠️ Could not set manual calculation: {e}")

        # === STEP 10: Aggressive calculation if needed (from Jupyter logic) ===
        if abs(base_target_value - original_forecast_value) < 1e-10:
            capture_print("\n🔄 Target still unchanged, trying more aggressive calculation...")
            try:
                # Try calculating specific sheets
                wb.sheets[target_sheet].calculate()
                time.sleep(1)
                app.calculate()
                time.sleep(2)
                
                # Check again
                final_check = wb.sheets[target_sheet].range(target_cell).value
                capture_print(f"📌 Target after aggressive calc: {final_check:.6f}")
                
                if abs(final_check - original_forecast_value) > 1e-10:
                    base_target_value = final_check
                    capture_print("✅ Aggressive calculation worked!")
                else:
                    capture_print("⚠️ Target still unchanged - there may be an issue with the model or inputs")
                    
            except Exception as e:
                capture_print(f"⚠️ Error during aggressive calculation: {e}")

        # === STEP 11: Group baseline inputs (COMPLETE Jupyter logic) ===
        def extract_row_from_cell(cell):
            """Extract row number from Excel cell reference (e.g., 'A5' -> 5, 'BC123' -> 123)"""
            match = re.search(r'(\d+)', cell)
            return int(match.group(1)) if match else None

        def find_common_description_part(descriptions):
            """Find the common part among multiple descriptions"""
            if not descriptions:
                return ""
            if len(descriptions) == 1:
                return descriptions[0]
            
            # Find common prefix
            common_prefix = ""
            min_len = min(len(desc) for desc in descriptions)
            
            for i in range(min_len):
                chars = set(desc[i] for desc in descriptions)
                if len(chars) == 1:
                    common_prefix += list(chars)[0]
                else:
                    break
            
            # Find common suffix
            common_suffix = ""
            for i in range(1, min_len - len(common_prefix) + 1):
                chars = set(desc[-i] for desc in descriptions)
                if len(chars) == 1:
                    common_suffix = list(chars)[0] + common_suffix
                else:
                    break
            
            # Combine and clean up
            common_part = (common_prefix + common_suffix).strip()
            
            # If common part is too short or empty, use the first description
            if len(common_part) < 10:
                return descriptions[0]
            
            return common_part

        # Group by sheet and row instead of by enriched description
        grouped_inputs = defaultdict(list)
        for item in baseline_inputs:
            sheet = item["sheet"]
            cell = item["cell"]
            row = extract_row_from_cell(cell)
            
            if row is not None:
                # Group by sheet and row
                group_key = f"{sheet}_row_{row}"
                grouped_inputs[group_key].append(item)
            else:
                # Fallback: group by individual cell if row extraction fails
                group_key = f"{sheet}_{cell}"
                grouped_inputs[group_key].append(item)

        capture_print(f"\n📊 Grouped into {len(grouped_inputs)} categories (by sheet and row)")

        # === Create final grouped results with common descriptions ===
        final_grouped_inputs = {}
        for group_key, items in grouped_inputs.items():
            # Extract all descriptions from items in this group
            descriptions = [item['desc'] for item in items]
            
            # Find common description part
            common_desc = find_common_description_part(descriptions)
            
            # Create a more readable group name
            if len(items) == 1:
                final_key = common_desc
            else:
                sheet_name = items[0]['sheet']
                row_num = extract_row_from_cell(items[0]['cell'])
                final_key = f"{common_desc} (Row {row_num})"
            
            final_grouped_inputs[final_key] = items

        # === Show grouping summary ===
        capture_print("\n📋 Grouping Summary:")
        for i, (desc, items) in enumerate(final_grouped_inputs.items(), 1):
            cells_in_group = [f"{item['sheet']}!{item['cell']}" for item in items]
            cells_display = ", ".join(cells_in_group[:3])
            if len(cells_in_group) > 3:
                cells_display += f" (+{len(cells_in_group)-3} more)"
            capture_print(f"  {i:2d}. {desc:<50} | {len(items):2d} cells | {cells_display}")

        # === Diagnostic: Check for ungrouped inputs ===
        grouped_cells = {f"{item['sheet']}!{item['cell']}" for group in final_grouped_inputs.values() for item in group}
        input_cells = {f"{item['sheet']}!{item['cell']}" for item in baseline_inputs}
        missing_cells = input_cells - grouped_cells

        if missing_cells:
            capture_print(f"\n⚠️ {len(missing_cells)} baseline input cells were not grouped:")
            for cell in sorted(list(missing_cells)[:5]):  # Show first 5
                capture_print(f"  - {cell}")
            if len(missing_cells) > 5:
                capture_print(f"  ... and {len(missing_cells) - 5} more")
        else:
            capture_print("\n✅ All baseline inputs were grouped correctly.")

        # === STEP 12: COMPLETE Attribution (change-back by group) - ALL GROUPS ===
        results = []
        total_groups = len(final_grouped_inputs)

        for group_idx, (group_desc, items) in enumerate(final_grouped_inputs.items(), 1):
            capture_print(f"\n🔄 [{group_idx}/{total_groups}] Changing back group: {group_desc}")

            target_before = wb.sheets[target_sheet].range(target_cell).value

            # Revert each cell in this group to its original forecast value
            for item in items:
                sheet, cell = item["sheet"], item["cell"]
                backup_key = f"{sheet}!{cell}"
                
                if backup_key in original_values_backup:
                    backup_data = original_values_backup[backup_key]
                    original_val = backup_data["value"]
                    original_formula = backup_data["formula"]
                    
                    try:
                        rng = wb.sheets[sheet].range(cell)
                        
                        # Check what type of original content this cell had
                        if original_formula and isinstance(original_formula, str) and original_formula.startswith('='):
                            # Cell originally had a formula like "=SUM(A1:A10)"
                            rng.formula = original_formula
                        else:
                            # Cell originally had just a number (no formula)
                            rng.formula = None  # Clear any temporary formula
                            rng.value = original_val
                            
                    except Exception as e:
                        capture_print(f"⚠️ Error reverting {sheet}!{cell}: {e}")

            # Calculate after reverting this group
            for j in range(2):  # Reduced attempts for speed
                try:
                    app.calculate()
                    time.sleep(1)
                    break
                except Exception as e:
                    capture_print(f"⚠️ Excel busy (attempt {j+1}): {e}")
                    time.sleep(2)

            new_target_value = wb.sheets[target_sheet].range(target_cell).value
            if new_target_value is not None:
                delta = new_target_value - base_target_value
                capture_print(f"   🎯 Target: {target_before:.4f} → {new_target_value:.4f} (Δ = {delta:+.4f})")
            else:
                delta = 0
                capture_print(f"   🎯 Target became None - using delta = 0")

            results.append({
                "description": group_desc,  # Use the common description for the row group
                "cells": [f"{item['sheet']}!{item['cell']}" for item in items],
                "delta": delta
            })

            # Restore cells in this group back to baseline
            for item in items:
                try:
                    rng = wb.sheets[item["sheet"]].range(item["cell"])
                    rng.value = item["change_to"]
                except Exception as e:
                    capture_print(f"⚠️ Error restoring {item['sheet']}!{item['cell']}: {e}")

        # === STEP 13: Normalize results (from Jupyter logic) ===
        total_delta = original_forecast_value - base_target_value

        capture_print(f"\n✅ Total groups analyzed: {len(results)}")
        capture_print(f"📉 Total Δ Target: {total_delta:+.4f}")

        for r in results:
            if abs(total_delta) > 1e-6:
                r["normalized_pct"] = (r["delta"] / total_delta) * 100
            else:
                r["normalized_pct"] = 0.0

        # === Add residual if needed ===
        sum_of_group_deltas = sum(r["delta"] for r in results)
        residual_delta = total_delta - sum_of_group_deltas

        if abs(residual_delta) > 1e-6:
            residual_pct = residual_delta / total_delta * 100 if abs(total_delta) > 1e-6 else 0
            results.append({
                "description": "Interrelated effects",
                "cells": [],
                "delta": residual_delta,
                "normalized_pct": residual_pct
            })
            capture_print(f"\n⚠️ Added Interrelated effects: Δ={residual_delta:+.4f}, % of Total={residual_pct:+.2f}%")

        # === STEP 14: Final report (COMPLETE Jupyter format) ===
        capture_print("\n📊 Final Attribution Results (% of total Δ Target):")
        capture_print(f"{'Row-Based Group Description':<60} | {'% of Total':>10} | {'Δ Target':>12} | {'Cells'}")
        capture_print("=" * 120)

        for r in sorted(results, key=lambda x: abs(x['delta']), reverse=True):
            cell_list = ", ".join(r['cells'][:3])  # Show first 3 cells
            if len(r['cells']) > 3:
                cell_list += f" (+{len(r['cells'])-3} more)"
            
            capture_print(
                f"{r['description']:<60} | {r['normalized_pct']:+8.2f}% | "
                f"{r['delta']:+10.4f} | {cell_list}"
            )

        # === STEP 15: Check sum (from Jupyter logic) ===
        capture_print(f"\n✅ Sum of all deltas: {sum_of_group_deltas:+.4f}")
        capture_print(f"📊 Total change: {total_delta:+.4f}")
        capture_print(f"🎯 Residual: {residual_delta:+.4f}")

        capture_print(f"\n🎉 Attribution analysis complete!")

        # Create summary
        summary = {
            "total_delta": total_delta,
            "sum_deltas": sum_of_group_deltas,
            "residual": residual_delta,
            "original_value": original_forecast_value,
            "baseline_value": base_target_value,
            "groups_analyzed": len(results),
            "changes_applied": changes_applied,
            "changes_verified": changes_verified
        }

        return {
            "results": results,
            "output": attribution_output,
            "summary": summary
        }

    except Exception as e:
        error_msg = f"❌ Error in attribution logic: {str(e)}"
        capture_print(error_msg)
        capture_print(f"Traceback: {traceback.format_exc()}")
        
        return {
            "results": [],
            "output": attribution_output,
            "error": error_msg,
            "summary": {}
        }
        
    finally:
        # === CRITICAL: Restore all original values AND FORMULAS ===
        capture_print("\n🔄 RESTORING ALL ORIGINAL VALUES AND FORMULAS...")
        
        restored_count = 0
        formula_restored = 0
        value_restored = 0
        
        for cell_key, backup_data in original_values_backup.items():
            try:
                sheet_name, cell_ref = cell_key.split('!')
                rng = wb.sheets[sheet_name].range(cell_ref)
                
                original_value = backup_data["value"]
                original_formula = backup_data["formula"]
                
                # Determine what type of content to restore
                if original_formula and isinstance(original_formula, str) and original_formula.startswith('='):
                    # This cell originally had a formula like "=SUM(A1:A10)" or "=A1*0.95"
                    rng.formula = original_formula
                    formula_restored += 1
                    
                    # Debug: Show formula restoration for first few cells
                    if formula_restored <= 3:
                        capture_print(f"   Restored formula to {cell_key}: '{original_formula}'")
                        
                elif original_formula == "" or original_formula is None:
                    # This cell originally had just a number (no formula)
                    rng.formula = None  # Ensure no formula
                    rng.value = original_value
                    value_restored += 1
                    
                    # Debug: Show value restoration for first few cells
                    if value_restored <= 3:
                        capture_print(f"   Restored value to {cell_key}: {original_value}")
                        
                else:
                    # Fallback: treat as value
                    rng.formula = None
                    rng.value = original_value
                    value_restored += 1
                    
                restored_count += 1
                
            except Exception as e:
                capture_print(f"⚠️ Error restoring {cell_key}: {e}")

        capture_print(f"✅ Restored {restored_count}/{len(original_values_backup)} cells")
        capture_print(f"   📝 Formulas restored: {formula_restored}")
        capture_print(f"   🔢 Values restored: {value_restored}")
        
        # Final calculation to ensure model is back to original state
        try:
            app.calculate()
            final_target = wb.sheets[target_sheet].range(target_cell).value
            if final_target is not None:
                capture_print(f"📌 Final target value: {final_target:.6f} (should match original: {original_forecast_value:.6f})")
            
            # Restore Excel settings
            if app:
                app.screen_updating = True
                app.enable_events = True
                app.display_alerts = True
                
        except Exception as e:
            capture_print(f"⚠️ Error in final restoration: {e}")

        capture_print("🧹 Cleanup completed - original file restored")




























def main():
    # Capture all output for debugging
    debug_output = []
    
    def debug_print(msg):
        print(msg)
        debug_output.append(msg)


    def print_filtered_trace(filtered_trace):
        debug_print("\n--- Filtered Forecast E Cells ---\n")
        for item in filtered_trace:
            period = item["period"]
            sheet = item['sheet']
            cell = item['cell']
            label = f"{item.get('description', '')} ({sheet}!{cell})"
            value = item.get('value')
            depth = item.get('depth')
            formula = item.get('formula') or ''
            constants = item.get("constants", [])
            const_str = ", ".join(f"{c:.4f}" for c in constants) if constants else ""
            external_ref = item.get("external_ref_segment", "")
            value_str = f"{value:.4f}" if isinstance(value, (float, int)) else str(value)

            debug_print(
                f"{label:<50} | Period: {period:<6} | Formula: {formula:<50} | "
                f"Value: {value_str:<12} | Depth: {depth} | External Ref: {external_ref} | Constants: {const_str}"
            )


    def print_terminal_inputs(filtered_trace, trace, all_col_to_period):
        debug_print("\n--- Enhanced Terminal Input Cells (FIXED) ---\n")
        included_count = 0
        excluded_count = 0

        for item in filtered_trace:
            if not item["period"].endswith("E"):
                continue
            if not isinstance(item.get("value"), (int, float)):
                continue
            if item.get("value") in (0, 0.0):
                continue

            period = item["period"]
            sheet = item['sheet']
            cell = item['cell']
            label = f"{item.get('description', '')} ({sheet}!{cell})"
            value = item.get('value')
            depth = item.get('depth')
            formula = item.get('formula') or ''
            constants = item.get("constants", [])
            const_str = ", ".join(f"{c:.4f}" for c in constants) if constants else ""
            external_ref = item.get("external_ref_segment", "")
            value_str = f"{value:.4f}" if isinstance(value, (float, int)) else str(value)

            is_included = is_terminal_input(item, trace, all_col_to_period)

            if is_included:
                status = "✅ INCLUDED"
                included_count += 1
                if not formula or "[" in formula:
                    reason = "No formula/External ref"
                elif re.match(r"^=[0-9.\-+*/ ()]+$", formula):
                    reason = "Constants only"
                elif constants:
                    excluded_constants = {0.25, 0.33, 0.5, 0.67, 0.75, 1/4, 1/3, 1/2, 2/3, 3/4,
                                        1, -1, 2, 3, 4, 52, 90, 100, 180, 270, 365, 1000, 10000,
                                        100000, 1000000}
                    meaningful_constants = [c for c in constants if c not in excluded_constants]
                    reason = f"Meaningful constants: {meaningful_constants}" if meaningful_constants else "Historical refs only"
                else:
                    reason = "Historical refs only"
            else:
                status = "❌ EXCLUDED"
                excluded_count += 1
                if constants:
                    excluded_constants = {0.25, 0.33, 0.5, 0.67, 0.75, 1/4, 1/3, 1/2, 2/3, 3/4,
                                        1, -1, 2, 3, 4, 52, 90, 100, 180, 270, 365, 1000, 10000,
                                        100000, 1000000}
                    meaningful_constants = [c for c in constants if c not in excluded_constants]
                    if meaningful_constants:
                        formula_upper = formula.upper()
                        for keyword in ["IF", "VLOOKUP", "MATCH", "OFFSET", "INDIRECT"]:
                            if keyword in formula_upper:
                                reason = f"Complex formula ({keyword})"
                                break
                        else:
                            reason = "References forecast periods"
                    else:
                        reason = "References forecast periods"
                else:
                    reason = "References forecast periods"

            debug_print(
                f"{status} | {label:<50} | Period: {period:<6} | Formula: {formula:<50} | "
                f"Value: {value_str:<12} | Constants: {const_str:<20} | Reason: {reason}"
            )

        debug_print(f"\n📊 Summary: {included_count} included, {excluded_count} excluded")
        debug_print("🔍 Review the above list to ensure business-meaningful constants are captured correctly!")

    def run_enrichment_for_all_terminals(terminal_inputs, trace, all_col_to_period):
        """
        Enrich and print descriptions for all terminal input cells with chain printing.
        """
        enriched_results = []

        debug_print("\n🔍 Enriching Terminal Inputs (Period-Matched Nearest Parents)\n")

        for i, item in enumerate(terminal_inputs, 1):
            debug_print(f"\n[{i}/{len(terminal_inputs)}] Processing: {item['sheet']}!{item['cell']}")
            debug_print(f"Original description: '{item.get('description', '')}' | Period: {item['period']} | Value: {item.get('value')}")

            enriched_desc, chain_info, parents = get_enriched_terminal_description(item, trace, all_col_to_period)

            if chain_info:
                debug_print(f"DEBUG CHAIN: Full chain length {chain_info['full_length']}: {'  →  '.join(chain_info['full_chain'])}")
                debug_print(f"DEBUG CHAIN: Filtered chain length {chain_info['filtered_length']} (same period only): {'  →  '.join(chain_info['filtered_chain'])}")
                if parents:
                    debug_print(f"DEBUG CHAIN: Selected parents: {parents}")
            else:
                debug_print("DEBUG CHAIN: No chain found")

            enriched_results.append({
                "sheet": item["sheet"],
                "cell": item["cell"],
                "period": item["period"],
                "value": item.get("value"),
                "depth": item.get("depth"),
                "original_desc": item.get("description", ""),
                "enriched_desc": enriched_desc,
                "chain_info": chain_info
            })

            debug_print(f"✅ Enriched: '{enriched_desc}'")

        debug_print("\n" + "="*100)
        debug_print("SUMMARY: Terminal Cell Enrichment (Nearest Parents in Same Period)")
        debug_print("="*100)

        enriched_count = 0
        for result in enriched_results:
            enriched = result["enriched_desc"]
            original = result["original_desc"]
            enriched_flag = "✅ ENRICHED" if ("→" in enriched and enriched != original) else "❌ NO ENRICHMENT"
            if enriched_flag == "✅ ENRICHED":
                enriched_count += 1

            chain_length = result["chain_info"]["filtered_length"] if result["chain_info"] else 0

            debug_print(f"{enriched_flag} | {result['sheet']}!{result['cell']:<10} | {result['period']} | Depth: {result['depth']} | Chain: {chain_length} | '{enriched}'")

        total = len(enriched_results)
        debug_print(f"\n📊 {enriched_count}/{total} enriched ({(enriched_count/total)*100:.1f}%)")

        return enriched_results
    



    def run_baseline_forcing(terminal_inputs, enriched_results, all_col_to_period, wb_values, wb_formulas):
        enriched_descriptions = {
            f"{result['sheet']}!{result['cell']}": result['enriched_desc']
            for result in enriched_results
        }

        sheet_names = {item['sheet'] for item in terminal_inputs}

        manual_today_date = datetime.date(2025, 4, 1)
        sheet_to_col_period = {}

        for sheet in sheet_names:
            if sheet in all_col_to_period:
                sheet_to_col_period[sheet] = all_col_to_period[sheet]
            else:
                sheet_to_col_period[sheet] = build_column_period_map(
                    wb_values, sheet, start_row=1, today_date=manual_today_date, max_row=150
                )

        terminal_period_map = {}
        for item in terminal_inputs:
            sheet = item["sheet"]
            cell = item["cell"]
            period = item["period"]
            col = re.match(r"[A-Z]+", cell).group(0)
            terminal_period_map.setdefault(sheet, {})[col] = period

        baseline_inputs = []
        recorded_cells = set()

        for item in terminal_inputs:
            try:
                # --- simplified to skip inner debug comments ---
                sheet = item["sheet"]
                cell = item["cell"]
                row = int(re.match(r"[A-Z]+(\d+)", cell).group(1))
                desc = item.get("description", "").lower()
                col_period_map = sheet_to_col_period[sheet]

                hist_cols = [col for col, period in col_period_map.items() if not period.endswith("E")]
                forecast_cols = [col for col, period in col_period_map.items() if period.endswith("E")]
                hist_cols_to_use = hist_cols

                if forecast_cols:
                    first_forecast_col = sorted(forecast_cols, key=column_index_from_string)[0]
                    first_index = column_index_from_string(first_forecast_col)
                    hist_cols_to_use = [
                        col for col in hist_cols if column_index_from_string(col) < first_index
                    ] or hist_cols

                terminal_period = item["period"]
                if terminal_period and re.match(r'^\d{4}E$', terminal_period):
                    year = int(terminal_period.replace('E', ''))
                    prior_year_cols = [col for col, period in col_period_map.items() if period == str(year - 1)]
                    latest_hist_col = prior_year_cols[0] if prior_year_cols else sorted(hist_cols_to_use, key=column_index_from_string)[-1]
                else:
                    latest_hist_col = sorted(hist_cols_to_use, key=column_index_from_string)[-1]

                latest_hist_cell = f"{latest_hist_col}{row}"
                latest_hist_val = wb_values[sheet][latest_hist_cell].value
                original_val = wb_values[sheet][cell].value
                number_format = wb_values[sheet][cell].number_format

                if should_force_zero(original_val, item.get("description", ""), number_format):
                    baseline_val, baseline_source = 0, "forced_zero_due_to_indicator"
                else:
                    formula = item.get("formula") or ""
                    ref_cells = re.findall(r"(?:'[^']+'|[^'!]+)![A-Z]+\d+|\$?[A-Z]+\$?\d+", formula)
                    ref_cells = [ref.replace("$", "") for ref in ref_cells]

                    if len(ref_cells) == 1:
                        ref = ref_cells[0]
                        ref_sheet, ref_cell = parse_sheet_reference(ref)
                        if ref_sheet is None:
                            ref_sheet = sheet
                        ref_period = lookup_column_period_from_map(all_col_to_period, ref_sheet, ref_cell)
                        baseline_val = (
                            wb_values[ref_sheet][ref_cell].value
                            if not ref_period.endswith("E")
                            else latest_hist_val
                        )
                        baseline_source = f"{ref_sheet}!{ref_cell}"
                    else:
                        baseline_val, baseline_source = latest_hist_val, latest_hist_cell

                enriched_desc = enriched_descriptions.get(f"{sheet}!{cell}", item.get("description", ""))

                # Build record
                formula_obj = wb_formulas[sheet][cell]
                cell_formula = formula_obj.value if isinstance(formula_obj.value, str) and formula_obj.value.startswith("=") else ""
                period_display = item["period"]

                if (sheet, cell) in recorded_cells:
                    continue
                recorded_cells.add((sheet, cell))

                baseline_inputs.append({
                    "sheet": sheet,
                    "cell": cell,
                    "desc": enriched_desc,
                    "original_value": original_val,
                    "change_to": baseline_val,
                    "type": "baseline_forcing",
                    "info": f"Forced to {baseline_val} from {baseline_source}",
                    "period": period_display,
                    "formula": cell_formula
                })

            except Exception as e:
                debug_print(f"⚠️ Error processing baseline for {sheet}!{cell}: {e}")

        # --- Final print summary ---
        debug_print("\n📊 Planned Baseline Changes (Terminal):")
        debug_print("=" * 150)
        for bi in baseline_inputs:
            orig_val_str = f"{bi['original_value']:.4f}" if isinstance(bi['original_value'], (int, float)) else str(bi['original_value'])
            change_to_str = f"{bi['change_to']:.4f}" if isinstance(bi['change_to'], (int, float)) else str(bi['change_to'])
            debug_print(f"Cell: {bi['sheet']}!{bi['cell']:<12} | Period: {bi['period']:<8}")
            debug_print(f"Description: {bi['desc']}")
            debug_print(f"Change: {orig_val_str} → {change_to_str} | Info: {bi['info']}")
            if bi.get("formula"):
                debug_print(f"Formula: {bi['formula']}")
            debug_print("-" * 100)

        debug_print(f"\n📊 Total baseline changes planned: {len(baseline_inputs)}")
        return baseline_inputs




    
    if len(sys.argv) < 7:
        debug_print("❌ Not enough arguments provided.")
        debug_print(f"Received {len(sys.argv)-1} arguments:")
        for i, arg in enumerate(sys.argv[1:], start=1):
            debug_print(f"  Arg {i}: {arg}")
        return

    # Get arguments from VBA (replace hardcoded values from Jupyter)
    workbook_path      = sys.argv[1]  # replaces WORKBOOK_FILE
    target_sheet       = sys.argv[2]  # replaces TARGET_SHEET
    target_variable    = sys.argv[3]
    target_period      = sys.argv[4]
    target_cell        = sys.argv[5]  # replaces TARGET_CELL
    baseline_comparison_cell      = sys.argv[6]  # replaces BASELINE_COMPARISON_CELL

    debug_print("✅ Arguments received successfully:")
    debug_print(f"  Workbook Path        : {workbook_path}")
    debug_print(f"  Target Sheet         : {target_sheet}")
    debug_print(f"  Target Variable      : {target_variable}")
    debug_print(f"  Target Period        : {target_period}")
    debug_print(f"  Target Cell          : {target_cell}")
    debug_print(f"  Actual Baseline Cell        : {baseline_comparison_cell}")

    # === HANDSHAKE ===
    signal_file = "C:/Users/pli/Desktop/PYTHON_READY.txt"
    with open(signal_file, "w", encoding="utf-8") as f:
        f.write("PYTHON_READY")
    debug_print("🚦 Signal sent to VBA - Python is ready")

    excel_closed_signal = "C:/Users/pli/Desktop/EXCEL_CLOSED.txt"
    debug_print("⏳ Waiting for VBA to close Excel...")
    
    timeout = 30
    start_time = time.time()
    
    while not os.path.exists(excel_closed_signal):
        time.sleep(0.5)
        if time.time() - start_time > timeout:
            debug_print("❌ Timeout waiting for Excel to close")
            return
    
    debug_print("✅ Excel closed signal received")
    
    # === CRITICAL: Wait a few more seconds for Excel to actually close ===
    debug_print("⏳ Waiting additional time for Excel to fully close...")
    time.sleep(5)  # Give Excel 5 seconds to actually quit
    debug_print("✅ Additional wait completed")
    
    # === EXACT JUPYTER NOTEBOOK LOGIC ===
    try:
        debug_print("📊 Loading workbooks...")
        # Your exact code: wb_formulas = load_workbook(WORKBOOK_FILE, data_only=False)
        wb_formulas = load_workbook(workbook_path, data_only=False)
        # Your exact code: wb_values = load_workbook(WORKBOOK_FILE, data_only=True)
        wb_values = load_workbook(workbook_path, data_only=True)
        debug_print("✅ Successfully loaded workbooks")
        
        # Your exact variable: manual_today_date = datetime.date(2025, 6, 24)
        manual_today_date = datetime.date(2025, 6, 24)
        
        debug_print(f"📋 Available sheets: {wb_values.sheetnames}")
        
        # Your exact function call:
        debug_print("🗓️ Building col_to_period maps for all sheets...")
        all_col_to_period = build_all_sheets_col_to_period(
            wb_values, 
            today_date=manual_today_date,
            max_row=150
        )

        # Your exact output format:
        debug_print("\n=== COLUMN TO PERIOD MAPPING RESULTS ===")
        if all_col_to_period:
            for sheet_name, col_to_period in all_col_to_period.items():
                debug_print(f"\n{sheet_name}:")
                for col, period in sorted(col_to_period.items()):
                    debug_print(f"  {col}: {period}")
        else:
            debug_print("No mappings found in any sheets")

        # Your exact test case:
        test_result = lookup_column_period_from_map(all_col_to_period, target_sheet, target_cell)
        debug_print(f"\nTest: {target_sheet}!{target_cell} period = '{test_result}'")
        
        if baseline_comparison_cell != "N/A":
            baseline_test = lookup_column_period_from_map(all_col_to_period, target_sheet, baseline_comparison_cell)
            debug_print(f"Test: {target_sheet}!{baseline_comparison_cell} period = '{baseline_test}'")
        
        # Additional debugging - show what's actually in target cells
        if target_sheet in wb_values.sheetnames:
            sheet = wb_values[target_sheet]
            debug_print(f"\n🔍 Debug: Sheet '{target_sheet}' has {sheet.max_row} rows, {sheet.max_column} columns")
            
            # Show target cell area
            target_col = re.match(r"([A-Z]+)", target_cell).group(1)
            debug_print(f"🔍 Debug: Looking at column {target_col} (first 20 rows):")
            for row in range(1, 21):
                cell_value = sheet[f"{target_col}{row}"].value
                if cell_value is not None:
                    debug_print(f"  {target_col}{row}: '{cell_value}' ({type(cell_value).__name__})")
            
            # Test regex on actual data
            debug_print(f"🔍 Debug: Testing regex pattern on sample data...")
            pattern = re.compile(r'\b(\d{1,2}Q\d{2,4}E?|\d{4}E?|FY\d{2,4}E?)\b', re.I)
            for row in range(1, 21):
                cell_value = sheet[f"{target_col}{row}"].value
                if cell_value and isinstance(cell_value, str):
                    match = pattern.search(cell_value.strip())
                    if match:
                        debug_print(f"  REGEX MATCH: {target_col}{row} '{cell_value}' -> '{match.group(1)}'")
                elif cell_value and isinstance(cell_value, (int, float)) and 2000 <= cell_value <= 2050:
                    debug_print(f"  YEAR FOUND: {target_col}{row} = {cell_value}")

            # === PROCESS TARGET CELLS ===
            if target_sheet in wb_formulas.sheetnames:
                debug_print(f"\n📋 Processing target sheet: {target_sheet}")
                
                # Get target cell data
                target_formula = wb_formulas[target_sheet][target_cell].value
                target_value = wb_values[target_sheet][target_cell].value
                target_period_found = lookup_column_period_from_map(all_col_to_period, target_sheet, target_cell)
                
                debug_print(f"Target Cell {target_cell}:")
                debug_print(f"   Formula: {target_formula}")
                debug_print(f"   Value: {target_value}")
                debug_print(f"   Period: {target_period_found}")
                
                # Get row number for label resolution
                target_row = int(re.search(r'(\d+)', target_cell).group(1))
                target_label = resolve_label_with_hierarchy(wb_values, target_sheet, target_row)
                debug_print(f"   Label: {target_label}")
                
                # Extract numeric constants
                if target_formula and isinstance(target_formula, str):
                    constants = extract_numeric_constants(target_formula)
                    if constants:
                        debug_print(f"   Numeric constants: {constants}")
                
                # === PERFORM FORMULA TRACING ===
                debug_print(f"\n🔍 Performing formula trace for {target_cell}...")
                
                trace = trace_merged(wb_values, wb_formulas, target_sheet, target_cell, all_col_to_period, max_depth=30)
                debug_print(f"\n=== FORMULA TRACE RESULTS ===")
                debug_print(f"Found {len(trace)} trace items:")

                ############################ --- Filter and Print E-period, non-zero trace items ---
                filtered_trace = [
                    item for item in trace
                    if item["period"].endswith("E")
                    and not is_effectively_zero(item.get("value"))
                    and not is_error_value(item.get("value"))
                ]
                terminal_inputs = [
                    item for item in filtered_trace
                    if is_terminal_input(item, trace, all_col_to_period)
                    and isinstance(item.get("value"), (int, float))
                    and item.get("value") not in (0, 0.0)
                ]

                print_filtered_trace(filtered_trace)
                print_terminal_inputs(filtered_trace, trace, all_col_to_period)
                enriched_results = run_enrichment_for_all_terminals(terminal_inputs, trace, all_col_to_period)
                baseline_inputs = run_baseline_forcing(terminal_inputs, enriched_results, all_col_to_period, wb_values, wb_formulas)

                
                # Add this section right after your baseline_inputs are created
                # Replace the complex attribution call with this simple verification:
                
                
                                

                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===

                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===

                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===

                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===


                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===

                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===

                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===

                debug_print(f"\n🚀 Starting COMPLETE attribution analysis with {len(baseline_inputs)} baseline inputs...")

                attribution_results = run_attribution_analysis_complete(
                    baseline_inputs=baseline_inputs,
                    workbook_path=workbook_path,
                    target_sheet=target_sheet,
                    target_cell=target_cell
                )

                # Then continue with your existing result processing code...

                # Then continue with your existing result processing code...

                # Then continue with your existing result processing code...

                # Then continue with your existing result processing code...


                # Add attribution output to debug_output for file writing
                if attribution_results and "output" in attribution_results:
                    debug_print("\n" + "="*60)
                    debug_print("🎯 ATTRIBUTION ANALYSIS RESULTS")
                    debug_print("="*60)
                    
                    for line in attribution_results["output"]:
                        debug_print(line)
                    
                    # Add summary to debug output
                    if "summary" in attribution_results and attribution_results["summary"]:
                        summary = attribution_results["summary"]
                        debug_print(f"\n📈 ATTRIBUTION SUMMARY:")
                        debug_print(f"   Original Value: {summary.get('original_value', 'N/A')}")
                        debug_print(f"   Baseline Value: {summary.get('baseline_value', 'N/A')}")
                        debug_print(f"   Total Change:   {summary.get('total_delta', 'N/A')}")
                        debug_print(f"   Sum of Groups:  {summary.get('sum_deltas', 'N/A')}")
                        debug_print(f"   Residual:       {summary.get('residual', 'N/A')}")
                        debug_print(f"   Groups:         {summary.get('groups_analyzed', 'N/A')}")
                        debug_print(f"   Changes Applied: {summary.get('changes_applied', 'N/A')}")
                        debug_print(f"   Changes Failed:  {summary.get('changes_failed', 'N/A')}")
                        
                    # Check for errors
                    if "error" in attribution_results:
                        debug_print(f"\n❌ ATTRIBUTION ERROR: {attribution_results['error']}")
                else:
                    debug_print("⚠️ No attribution results returned or invalid format")

                debug_print("✅ Attribution analysis section completed!")

                
                for item in trace:
                    sheet = item['sheet']
                    cell = item['cell']
                    cell_loc = f"{sheet}!{cell}"

                    # Extract value directly
                    cell_value = wb_values[sheet][cell].value
                    if isinstance(cell_value, str):
                        name = cell_value.strip()
                    else:
                        name = item.get('description') or ''

                    formula = item.get('formula') or ''
                    value = item.get('value')
                    depth = item.get('depth')
                    external_ref = item.get("external_ref_segment", "")

                    col = re.match(r"[A-Z]+", cell).group(0)
                    period = item["period"]

                    label = f"{name} ({cell_loc})" if name else cell_loc
                    value_str = f"{value:.4f}" if isinstance(value, (int, float)) else (value or "")
                    constants = item.get("constants", [])
                    const_str = ", ".join(f"{c:.4f}" for c in constants) if constants else ""

                    trace_line = (
                        f"{label:<50} | Period: {period:<6} | Formula: {formula:<50} | "
                        f"Value: {value_str:<12} | Depth: {depth} | External Ref: {external_ref} | Constants: {const_str}"
                    )
                    
                    debug_print(trace_line)


                
                # Process baseline cell if available
                if baseline_comparison_cell != "N/A":
                    baseline_formula = wb_formulas[target_sheet][baseline_comparison_cell].value
                    baseline_value = wb_values[target_sheet][baseline_comparison_cell].value
                    baseline_period_found = lookup_column_period_from_map(all_col_to_period, target_sheet, baseline_comparison_cell)
                    
                    debug_print(f"\nBaseline Cell {baseline_comparison_cell}:")
                    debug_print(f"   Formula: {baseline_formula}")
                    debug_print(f"   Value: {baseline_value}")
                    debug_print(f"   Period: {baseline_period_found}")
                    
                    # Calculate difference
                    if target_value is not None and baseline_value is not None:
                        try:
                            difference = target_value - baseline_value
                            if baseline_value != 0:
                                percent_change = (difference / baseline_value) * 100
                                debug_print(f"\n📊 Analysis Results:")
                                debug_print(f"   Difference: {difference:,.2f}")
                                debug_print(f"   Percent Change: {percent_change:.2f}%")
                            else:
                                debug_print(f"\n📊 Analysis Results:")
                                debug_print(f"   Difference: {difference:,.2f}")
                                debug_print(f"   Percent Change: N/A (baseline is zero)")
                        except (TypeError, ValueError):
                            debug_print("⚠️  Could not calculate numeric difference")
            
            wb_formulas.close()
            wb_values.close()
        
    except Exception as e:
        debug_print(f"❌ Error processing workbook: {str(e)}")
        import traceback
        debug_print(f"Traceback: {traceback.format_exc()}")

    # === WRITE COMPREHENSIVE RESULTS WITH ALL DEBUG OUTPUT ===
    try:
        result_file = "C:/Users/pli/Desktop/ARG_CONFIRMATION.txt"
        
        with open(result_file, "w", encoding="utf-8") as f:
            f.write("=== PYTHON ARGUMENT CONFIRMATION ===\n")
            f.write(f"Timestamp            : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Workbook Path        : {workbook_path}\n")
            f.write(f"Target Sheet         : {target_sheet}\n")
            f.write(f"Target Variable      : {target_variable}\n")
            f.write(f"Target Period        : {target_period}\n")
            f.write(f"Target Cell          : {target_cell}\n")
            f.write(f"Baseline Cell        : {baseline_comparison_cell}\n")
            f.write("\n=== FULL DEBUG OUTPUT ===\n")
            
            # Write all debug output
            for line in debug_output:
                f.write(f"{line}\n")
            
            f.write("\n=== ANALYSIS COMPLETED ===\n")
        
        debug_print(f"✅ Results written to {result_file}")
    except Exception as e:
        debug_print(f"❌ Error writing results: {str(e)}")
    
    # Clean up
    try:
        if os.path.exists(signal_file):
            os.remove(signal_file)
        if os.path.exists(excel_closed_signal):
            os.remove(excel_closed_signal)
    except:
        pass
    
    debug_print("🎉 Analysis completed!")

if __name__ == "__main__":
    main()








def main():
    # Capture all output for debugging
    debug_output = []
    
    def debug_print(msg):
        print(msg)
        debug_output.append(msg)

    def run_enrichment_for_all_terminals(terminal_inputs, trace, all_col_to_period):
        """
        Enrich and print descriptions for all terminal input cells with chain printing.
        """
        enriched_results = []

        for i, item in enumerate(terminal_inputs, 1):

            enriched_desc, chain_info, parents = get_enriched_terminal_description(item, trace, all_col_to_period)

            enriched_results.append({
                "sheet": item["sheet"],
                "cell": item["cell"],
                "period": item["period"],
                "value": item.get("value"),
                "depth": item.get("depth"),
                "original_desc": item.get("description", ""),
                "enriched_desc": enriched_desc,
                "chain_info": chain_info
            })

        enriched_count = 0
        for result in enriched_results:
            enriched = result["enriched_desc"]
            original = result["original_desc"]
            enriched_flag = "✅ ENRICHED" if ("→" in enriched and enriched != original) else "❌ NO ENRICHMENT"
            if enriched_flag == "✅ ENRICHED":
                enriched_count += 1

            chain_length = result["chain_info"]["filtered_length"] if result["chain_info"] else 0
        total = len(enriched_results)

        return enriched_results
    

    def run_baseline_forcing(terminal_inputs, enriched_results, all_col_to_period, wb_values, wb_formulas):
        enriched_descriptions = {
            f"{result['sheet']}!{result['cell']}": result['enriched_desc']
            for result in enriched_results
        }

        sheet_names = {item['sheet'] for item in terminal_inputs}

        manual_today_date = datetime.date(2025, 4, 1)
        sheet_to_col_period = {}

        for sheet in sheet_names:
            if sheet in all_col_to_period:
                sheet_to_col_period[sheet] = all_col_to_period[sheet]
            else:
                sheet_to_col_period[sheet] = build_column_period_map(
                    wb_values, sheet, start_row=1, today_date=manual_today_date, max_row=150
                )

        terminal_period_map = {}
        for item in terminal_inputs:
            sheet = item["sheet"]
            cell = item["cell"]
            period = item["period"]
            col = re.match(r"[A-Z]+", cell).group(0)
            terminal_period_map.setdefault(sheet, {})[col] = period

        baseline_inputs = []
        recorded_cells = set()

        for item in terminal_inputs:
            try:
                # --- simplified to skip inner debug comments ---
                sheet = item["sheet"]
                cell = item["cell"]
                row = int(re.match(r"[A-Z]+(\d+)", cell).group(1))
                desc = item.get("description", "").lower()
                col_period_map = sheet_to_col_period[sheet]

                hist_cols = [col for col, period in col_period_map.items() if not period.endswith("E")]
                forecast_cols = [col for col, period in col_period_map.items() if period.endswith("E")]
                hist_cols_to_use = hist_cols

                if forecast_cols:
                    first_forecast_col = sorted(forecast_cols, key=column_index_from_string)[0]
                    first_index = column_index_from_string(first_forecast_col)
                    hist_cols_to_use = [
                        col for col in hist_cols if column_index_from_string(col) < first_index
                    ] or hist_cols

                terminal_period = item["period"]
                if terminal_period and re.match(r'^\d{4}E$', terminal_period):
                    year = int(terminal_period.replace('E', ''))
                    prior_year_cols = [col for col, period in col_period_map.items() if period == str(year - 1)]
                    latest_hist_col = prior_year_cols[0] if prior_year_cols else sorted(hist_cols_to_use, key=column_index_from_string)[-1]
                else:
                    latest_hist_col = sorted(hist_cols_to_use, key=column_index_from_string)[-1]

                latest_hist_cell = f"{latest_hist_col}{row}"
                latest_hist_val = wb_values[sheet][latest_hist_cell].value
                original_val = wb_values[sheet][cell].value
                number_format = wb_values[sheet][cell].number_format

                if should_force_zero(original_val, item.get("description", ""), number_format):
                    baseline_val, baseline_source = 0, "forced_zero_due_to_indicator"
                else:
                    formula = item.get("formula") or ""
                    ref_cells = re.findall(r"(?:'[^']+'|[^'!]+)![A-Z]+\d+|\$?[A-Z]+\$?\d+", formula)
                    ref_cells = [ref.replace("$", "") for ref in ref_cells]

                    if len(ref_cells) == 1:
                        ref = ref_cells[0]
                        ref_sheet, ref_cell = parse_sheet_reference(ref)
                        if ref_sheet is None:
                            ref_sheet = sheet
                        ref_period = lookup_column_period_from_map(all_col_to_period, ref_sheet, ref_cell)
                        baseline_val = (
                            wb_values[ref_sheet][ref_cell].value
                            if not ref_period.endswith("E")
                            else latest_hist_val
                        )
                        baseline_source = f"{ref_sheet}!{ref_cell}"
                    else:
                        baseline_val, baseline_source = latest_hist_val, latest_hist_cell

                enriched_desc = enriched_descriptions.get(f"{sheet}!{cell}", item.get("description", ""))

                # Build record
                formula_obj = wb_formulas[sheet][cell]
                cell_formula = formula_obj.value if isinstance(formula_obj.value, str) and formula_obj.value.startswith("=") else ""
                period_display = item["period"]

                if (sheet, cell) in recorded_cells:
                    continue
                recorded_cells.add((sheet, cell))

                baseline_inputs.append({
                    "sheet": sheet,
                    "cell": cell,
                    "desc": enriched_desc,
                    "original_value": original_val,
                    "change_to": baseline_val,
                    "type": "baseline_forcing",
                    "info": f"Forced to {baseline_val} from {baseline_source}",
                    "period": period_display,
                    "formula": cell_formula
                })

            except Exception as e:
                continue
        return baseline_inputs


    if len(sys.argv) < 7:
        debug_print("❌ Not enough arguments provided.")
        debug_print(f"Received {len(sys.argv)-1} arguments:")
        for i, arg in enumerate(sys.argv[1:], start=1):
            debug_print(f"  Arg {i}: {arg}")
        return

    # Get arguments from VBA (replace hardcoded values from Jupyter)
    workbook_path      = sys.argv[1]  # replaces WORKBOOK_FILE
    target_sheet       = sys.argv[2]  # replaces TARGET_SHEET
    target_variable    = sys.argv[3]
    target_period      = sys.argv[4]
    target_cell        = sys.argv[5]  # replaces TARGET_CELL
    baseline_comparison_cell      = sys.argv[6]  # replaces BASELINE_COMPARISON_CELL

    # === HANDSHAKE ===
    signal_file = "C:/Users/pli/Desktop/PYTHON_READY.txt"
    with open(signal_file, "w", encoding="utf-8") as f:
        f.write("PYTHON_READY")
    debug_print("🚦 Signal sent to VBA - Python is ready")

    excel_closed_signal = "C:/Users/pli/Desktop/EXCEL_CLOSED.txt"
    debug_print("⏳ Waiting for VBA to close Excel...")
    
    timeout = 30
    start_time = time.time()
    
    while not os.path.exists(excel_closed_signal):
        time.sleep(0.5)
        if time.time() - start_time > timeout:
            debug_print("❌ Timeout waiting for Excel to close")
            return
    
    debug_print("✅ Excel closed signal received")
    
    # === CRITICAL: Wait a few more seconds for Excel to actually close ===
    debug_print("⏳ Waiting additional time for Excel to fully close...")
    time.sleep(5)  # Give Excel 5 seconds to actually quit
    debug_print("✅ Additional wait completed")
    
    # === EXACT JUPYTER NOTEBOOK LOGIC ===
    try:
        debug_print("📊 Loading workbooks...")
        # Your exact code: wb_formulas = load_workbook(WORKBOOK_FILE, data_only=False)
        wb_formulas = load_workbook(workbook_path, data_only=False)
        # Your exact code: wb_values = load_workbook(WORKBOOK_FILE, data_only=True)
        wb_values = load_workbook(workbook_path, data_only=True)
        debug_print("✅ Successfully loaded workbooks")
        
        # Your exact variable: manual_today_date = datetime.date(2025, 6, 24)
        manual_today_date = datetime.date(2025, 6, 24)
        
        
        # Your exact function call:
        debug_print("🗓️ Building col_to_period maps for all sheets...")
        all_col_to_period = build_all_sheets_col_to_period(
            wb_values, 
            today_date=manual_today_date,
            max_row=150
        )

        # Your exact test case:
        test_result = lookup_column_period_from_map(all_col_to_period, target_sheet, target_cell)

        # Additional debugging - show what's actually in target cells
        if target_sheet in wb_values.sheetnames:
            sheet = wb_values[target_sheet]
            
            # Show target cell area
            target_col = re.match(r"([A-Z]+)", target_cell).group(1)
            for row in range(1, 21):
                cell_value = sheet[f"{target_col}{row}"].value
                if cell_value is not None:
                    debug_print(f"  {target_col}{row}: '{cell_value}' ({type(cell_value).__name__})")
            

            # === PROCESS TARGET CELLS ===
            if target_sheet in wb_formulas.sheetnames:
                # Get target cell data
                target_formula = wb_formulas[target_sheet][target_cell].value
                target_value = wb_values[target_sheet][target_cell].value
                target_period_found = lookup_column_period_from_map(all_col_to_period, target_sheet, target_cell)
                
                
                # Get row number for label resolution
                target_row = int(re.search(r'(\d+)', target_cell).group(1))
                target_label = resolve_label_with_hierarchy(wb_values, target_sheet, target_row)
                
                # Extract numeric constants
                if target_formula and isinstance(target_formula, str):
                    constants = extract_numeric_constants(target_formula)
                
                trace = trace_merged(wb_values, wb_formulas, target_sheet, target_cell, all_col_to_period, max_depth=30)


                ############################ --- Filter and Print E-period, non-zero trace items ---
                filtered_trace = [
                    item for item in trace
                    if item["period"].endswith("E")
                    and not is_effectively_zero(item.get("value"))
                    and not is_error_value(item.get("value"))
                ]
                terminal_inputs = [
                    item for item in filtered_trace
                    if is_terminal_input(item, trace, all_col_to_period)
                    and isinstance(item.get("value"), (int, float))
                    and item.get("value") not in (0, 0.0)
                ]
                enriched_results = run_enrichment_for_all_terminals(terminal_inputs, trace, all_col_to_period)
                baseline_inputs = run_baseline_forcing(terminal_inputs, enriched_results, all_col_to_period, wb_values, wb_formulas)

                
                # Add this section right after your baseline_inputs are created
                # Replace the complex attribution call with this simple verification:
                
                
                                

                # === IN YOUR MAIN FUNCTION, REPLACE THE ATTRIBUTION CALL WITH: ===

                debug_print(f"\n🚀 Starting COMPLETE attribution analysis with {len(baseline_inputs)} baseline inputs...")

                attribution_results = run_attribution_analysis_complete(
                    baseline_inputs=baseline_inputs,
                    workbook_path=workbook_path,
                    target_sheet=target_sheet,
                    target_cell=target_cell
                )

                # Then continue with your existing result processing code...



                # Add attribution output to debug_output for file writing
                if attribution_results and "output" in attribution_results:
                    debug_print("\n" + "="*60)
                    debug_print("🎯 ATTRIBUTION ANALYSIS RESULTS")
                    debug_print("="*60)
                    
                    for line in attribution_results["output"]:
                        debug_print(line)
                    
                    # Add summary to debug output
                    if "summary" in attribution_results and attribution_results["summary"]:
                        summary = attribution_results["summary"]
                        debug_print(f"\n📈 ATTRIBUTION SUMMARY:")
                        debug_print(f"   Original Value: {summary.get('original_value', 'N/A')}")
                        debug_print(f"   Baseline Value: {summary.get('baseline_value', 'N/A')}")
                        debug_print(f"   Total Change:   {summary.get('total_delta', 'N/A')}")
                        debug_print(f"   Sum of Groups:  {summary.get('sum_deltas', 'N/A')}")
                        debug_print(f"   Residual:       {summary.get('residual', 'N/A')}")
                        debug_print(f"   Groups:         {summary.get('groups_analyzed', 'N/A')}")
                        debug_print(f"   Changes Applied: {summary.get('changes_applied', 'N/A')}")
                        debug_print(f"   Changes Failed:  {summary.get('changes_failed', 'N/A')}")
                        
                    # Check for errors
                    if "error" in attribution_results:
                        debug_print(f"\n❌ ATTRIBUTION ERROR: {attribution_results['error']}")
                else:
                    debug_print("⚠️ No attribution results returned or invalid format")

                debug_print("✅ Attribution analysis section completed!")


                
                # Process baseline cell if available
            wb_formulas.close()
            wb_values.close()
        
    except Exception as e:
        debug_print(f"❌ Error processing workbook: {str(e)}")
        import traceback
        debug_print(f"Traceback: {traceback.format_exc()}")

    # === WRITE COMPREHENSIVE RESULTS WITH ALL DEBUG OUTPUT ===
    try:
        result_file = "C:/Users/pli/Desktop/ARG_CONFIRMATION.txt"
        
        with open(result_file, "w", encoding="utf-8") as f:
            f.write("🎯 ATTRIBUTION ANALYSIS RESULTS\n")
            f.write("="*60 + "\n")
            for line in debug_output:
                f.write(f"{line}\n")

        
        debug_print(f"✅ Results written to {result_file}")
    except Exception as e:
        debug_print(f"❌ Error writing results: {str(e)}")
    
    # Clean up
    try:
        if os.path.exists(signal_file):
            os.remove(signal_file)
        if os.path.exists(excel_closed_signal):
            os.remove(excel_closed_signal)
    except:
        pass
    
    debug_print("🎉 Analysis completed!")

if __name__ == "__main__":
    main()