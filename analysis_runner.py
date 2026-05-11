# analysis_runner.py
from utils.excel_format import excel_output_format

from utils.workbook_utils import (
    build_all_sheets_col_to_period,
    lookup_column_period_from_map,
    resolve_label_with_hierarchy,
    parse_sheet_reference,
    build_column_period_map
)

from utils.trace_utils import (
    trace_merged,
    is_effectively_zero,
    is_error_value,
    is_terminal_input
)

from utils.enrichment_utils import (
    get_enriched_terminal_description,
    extract_numeric_constants, 
    run_baseline_forcing, 
    run_enrichment_for_all_terminals,
    should_force_zero
)

from utils.attribution_utils import (
    run_attribution_analysis_complete
)

import datetime
import re
from openpyxl import load_workbook
import requests
from openpyxl import load_workbook
from io import BytesIO

def run_analysis(
    workbook_path: str,
    target_sheet: str,
    target_variable: str,
    target_period: str,
    target_cell: str,
    baseline_comparison_cell: str
) -> dict:
    debug_output = []

    def debug_print(msg):
        print(msg, flush=True)
        debug_output.append(msg)

    from urllib.parse import urlparse

    def is_url(path: str) -> bool:
        try:
            result = urlparse(path)
            return all([result.scheme in ("http", "https"), result.netloc])
        except ValueError:
            return False



    try:
        debug_print("📊 Loading workbooks...")

        if is_url(workbook_path):
            debug_print("Loading workbook from URL.")

            # Download the file using requests
            response = requests.get(workbook_path)
            response.raise_for_status()  # Raises an error if the download failed

            # Load the workbook from the downloaded content
            wb_formulas = load_workbook(filename=BytesIO(response.content), data_only= False)
            wb_values = load_workbook(filename=BytesIO(response.content), data_only= True)
            debug_print("Workbook successful loaded from url")
        else:
            debug_print("Loading workbook from local path.")

            wb_formulas = load_workbook(workbook_path, data_only=False)
            wb_values = load_workbook(workbook_path, data_only=True)
            debug_print("Workbook successful loaded from local")




        manual_today_date = datetime.date(2025, 6, 24)

        debug_print("🗓️ Building col_to_period maps for all sheets...")
        all_col_to_period = build_all_sheets_col_to_period(
            wb_values,
            today_date=manual_today_date,
            max_row=150
        )

        target_formula = wb_formulas[target_sheet][target_cell].value
        target_value = wb_values[target_sheet][target_cell].value
        target_period_found = lookup_column_period_from_map(all_col_to_period, target_sheet, target_cell)

        target_row = int(re.search(r'(\d+)', target_cell).group(1))
        target_label = resolve_label_with_hierarchy(wb_values, target_sheet, target_row)

        constants = extract_numeric_constants(target_formula) if isinstance(target_formula, str) else []

        trace = trace_merged(wb_values, wb_formulas, target_sheet, target_cell, all_col_to_period, max_depth=30)

        filtered_trace = [
            item for item in trace
            if item["period"].endswith("E")
            and not is_effectively_zero(item.get("value"))
            and not is_error_value(item.get("value"))
        ]

        terminal_inputs = [
            item for item in filtered_trace
            if is_terminal_input(item, trace, all_col_to_period)
            and isinstance(item.get("value"), (int, float))
            and item.get("value") not in (0, 0.0)
        ]

        enriched_results = run_enrichment_for_all_terminals(terminal_inputs, trace, all_col_to_period)
        baseline_inputs = run_baseline_forcing(terminal_inputs, enriched_results, all_col_to_period, wb_values, wb_formulas)

        debug_print(f"\n🚀 Starting COMPLETE attribution analysis with {len(baseline_inputs)} baseline inputs...")

        # This function now handles file writing internally
        attribution_results = run_attribution_analysis_complete(
            baseline_inputs=baseline_inputs,
            workbook_path=workbook_path,
            target_sheet=target_sheet,
            target_cell=target_cell
        )

        output_lines = []
        if attribution_results and "output" in attribution_results:
            output_lines.extend(["="*60, "🎯 ATTRIBUTION ANALYSIS RESULTS", "="*60])
            output_lines.extend(attribution_results["output"])

            if "summary" in attribution_results and attribution_results["summary"]:
                summary = attribution_results["summary"]
                output_lines.append("\n📈 ATTRIBUTION SUMMARY:")
                output_lines.append(f"   Original Value: {summary.get('original_value', 'N/A')}")
                output_lines.append(f"   Baseline Value: {summary.get('baseline_value', 'N/A')}")
                output_lines.append(f"   Total Change:   {summary.get('total_delta', 'N/A')}")
                output_lines.append(f"   Sum of Groups:  {summary.get('sum_deltas', 'N/A')}")
                output_lines.append(f"   Residual:       {summary.get('residual', 'N/A')}")
                output_lines.append(f"   Groups:         {summary.get('groups_analyzed', 'N/A')}")
                output_lines.append(f"   Changes Applied: {summary.get('changes_applied', 'N/A')}")
                output_lines.append(f"   Changes Failed:  {summary.get('changes_failed', 'N/A')}")

        debug_print("✅ Attribution analysis section completed!")




        results = attribution_results['results']
        original_forecaset_value = attribution_results['original_target_value']
        base_target_value = attribution_results['baseline_target_value']

        excel_output_format(results = results, 
                            baseline_inputs = baseline_inputs,
                            WORKBOOK_FILE = workbook_path,
                            TARGET_SHEET = target_sheet,
                            TARGET_CELL = target_cell,
                            BASELINE_COMPARISON_CELL = baseline_comparison_cell,
                            original_forecast_value = original_forecaset_value,
                            base_target_value = base_target_value,
                            wb_values = wb_values
                            )


        wb_formulas.close()
        wb_values.close()


        return {
            "debug": debug_output,
            "output": output_lines,
            "summary": attribution_results.get("summary", {}) if attribution_results else {},
            "error": attribution_results.get("error") if attribution_results else None
        }

    except Exception as e:
        import traceback
        debug_print(f"❌ Error in run_analysis: {str(e)}")
        debug_print(traceback.format_exc())
        return {
            "debug": debug_output,
            "output": [],
            "summary": {},
            "error": str(e)
        }